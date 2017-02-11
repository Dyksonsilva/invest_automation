def relatorio(dt_base, cnpj, horario_bd):

    import pandas as pd
    import pymysql as db
    import numpy as np
    import math
    import openpyxl
    import datetime
    import logging
    from openpyxl.styles import Font, Color, Border, Side, PatternFill, Alignment
    from dependencias.Metodos.funcoes_auxiliares import full_path_from_database

    #Global Vars
    global id_relatorio_qo
    global header_id_carteira
    global header_nome
    global cnpj_fundo
    global administrador
    global gestor

    logger = logging.getLogger(__name__)
    end = full_path_from_database("get_output_var") + "relatorios/"

    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(end + 'relatorio_quadro_operacoes.xlsx', engine='xlsxwriter')
    workbook = writer.book
    numero_float = workbook.add_format({'num_format': '#,##0', 'bold': False})
    percent = workbook.add_format({'num_format': '0.0%', 'bold': False})
    contador_qo = 0
    contador_desp = 0
    contador_contraparte = 0
    contador_isin = 0
    contador_titulos = 0
    contador_grupo_prod = 0
    contador_acoes = 0
    contador_descasamento = ''

    logger.info("Conectando no Banco de dados")
    connection = db.connect('localhost', user='root', passwd='root', db='projeto_inv', use_unicode=True, charset="utf8")
    logger.info("Conexão com DB executada com sucesso")

    query = 'select * from projeto_inv.xml_header where cnpjcpf="' + cnpj + '" and dtposicao=' + '"' + dt_base + '";'
    df = pd.read_sql(query, con=connection)
    logger.info("Leitura do banco de dados executada com sucesso")

    if len(df) == 0:
        query = 'select * from projeto_inv.xml_header where cnpj="' + cnpj + '" and dtposicao=' + '"' + dt_base + '";'
        df = pd.read_sql(query, con=connection)
        logger.info("Leitura do banco de dados executada com sucesso")

    logger.info("Tratando dados")
    df = df.sort(['cnpj', 'cnpjcpf', 'data_bd'], ascending=[True, True, False])
    df = df.drop_duplicates(subset=['cnpj', 'cnpjcpf'], take_last=False)
    df = df.reset_index(level=None, drop=False, inplace=False, col_level=0, col_fill='')
    del df['index']


    header_id_carteira = df.get_value(0, 'header_id').astype(str)
    header_nome = df.get_value(0, 'nome')
    cnpj_fundo = cnpj
    administrador = df.get_value(0, 'nomeadm')
    gestor = df.get_value(0, 'nomegestor')

    # quadro de operaçoes
    query = 'select a.* from projeto_inv.xml_quadro_operacoes a right join (select header_id, max(data_bd) as data_bd from projeto_inv.xml_quadro_operacoes where header_id=' + header_id_carteira + ' group by 1) b on a.header_id=b.header_id and a.data_bd=b.data_bd;'
    qo = pd.read_sql(query, con=connection)

    if len(qo['cnpj'][qo.cnpj.notnull()]) > 0:
        qo['cnpj1'] = ""
        for i in range(0, len(qo)):
            if qo['cnpj'][i] > 0:
                qo['cnpj1'][i] = math.floor(qo['cnpj'][i])
        qo['cnpj2'] = qo['cnpj1'].astype(str).str.zfill(14)
        del qo['cnpj']
        del qo['cnpj1']
    qo = qo.rename(columns={'cnpj2': 'cnpj'})

    id_relatorio_qo = str(qo['id_relatorio_qo'][0])

    qo['perc_index'] = qo['perc_index'] / 100
    qo['tx_operacao'] = qo['tx_operacao'] / 100

    quadro_operacoes = qo.reindex(
        columns=['mtm_regra_xml', 'mtm_info', 'produto', 'dt_emissao', 'dt_compra', 'cnpj', 'contraparte', 'a_p_op',
                 'quantidade', 'dt_vencto', 'indexador', 'perc_index', 'tx_operacao', 'isin', 'ativo', 'caracteristica',
                 'cnpjfundo_1nivel', 'fundo', 'cnpjfundo_outros', 'fundo_ult_nivel'])
    quadro_operacoes = quadro_operacoes.rename(columns={'mtm_regra_xml': 'mtm_calculado'})

    if len(quadro_operacoes) != 0:
        aux_cora = pd.DataFrame(columns=['Dados Institucionais'])
        aux_cora.to_excel(writer, index=False, sheet_name='Dados Institucionais', startrow=1, startcol=1,
                          header=['Dados Institucionais'])

        quadro_operacoes.to_excel(writer, index=False, sheet_name='Quadro de Operações', startrow=3, startcol=1,
                                  header=['MtM Calculado (R$)', 'MtM Informado (R$)', 'Produto', 'Data de Emissão',
                                          'Data de Compra', 'CNPJ da Contraparte', 'Contraparte', 'Comprado/Vendido',
                                          'Quantidade', 'Data de Vencimento', 'Indexador', 'Percentual do índice (%)',
                                          'Taxa da Operação (%)', 'ISIN', 'Ativo', 'Característica',
                                          'CNPJ do Fundo - 1o Nível', 'Nome do Fundo - 1o Nível',
                                          'CNPJ do Fundo - último nível', 'Nome do Fundo - Último nível'])

        worksheet = writer.sheets['Quadro de Operações']
        worksheet.set_column('B:C', 12, numero_float)
        contador_qo = 1

    # Relatorio de despesas
    query = 'select a.* from projeto_inv.xml_quadro_despesas a right join (select header_id, max(data_bd) as data_bd from projeto_inv.xml_quadro_despesas where header_id=' + header_id_carteira + ' group by 1) b on a.header_id=b.header_id and a.data_bd=b.data_bd;'
    desp = pd.read_sql(query, con=connection)
    despfim = desp[['mtm_info', 'produto']]
    relatorio_desp = despfim.groupby(["produto"], as_index=False).sum()
    relatorio_desp = relatorio_desp.rename(columns={'mtm_info': 'Financeiro  (R$)', 'produto': 'Descrição'})

    if len(relatorio_desp) != 0:
        relatorio_desp.to_excel(writer, index=False, sheet_name='Despesas', startrow=3, startcol=1)
        worksheet = writer.sheets['Despesas']
        worksheet.set_column('C:C', 12, numero_float)
        contador_desp = 1

    # Relatório de rating por contraparte
    lista_credito = ['debênture', 'título privado', 'compromissada - debênture']
    credito = qo[qo['produto'].isin(lista_credito)]

    # Relatório de rating por contraparte
    base_contraparte = credito[['mtm_info', 'cnpj', 'contraparte']]
    contraparte = base_contraparte.groupby(['cnpj', 'contraparte'], as_index=False).sum()

    query = 'select distinct a.cnpj, a.agencia_tipo_rtg, a.rtg from projeto_inv.rating_contraparte as a right join (select max(data_bd) as data_bd from projeto_inv.rating_contraparte where dt_ref= "' + dt_base + '" ) as b on a.data_bd=b.data_bd;'
    rtg = pd.read_sql(query, con=connection)

    query = 'select cod_rtg, rtg from projeto_inv.de_para_rtg a right join (select max(data_bd) as data_bd from projeto_inv.de_para_rtg) b on a.data_bd = b.data_bd;'
    depara = pd.read_sql(query, con=connection)

    moodys = rtg[rtg['agencia_tipo_rtg'] == 'RTG_MDY_NSR_ISSUER']
    moodys = moodys.rename(columns={'rtg': "Moody's"})

    sp = rtg[rtg['agencia_tipo_rtg'] == 'RTG_SP_NATIONAL_LT_ISSUER_CREDIT']
    sp = sp.rename(columns={'rtg': "Standard & Poor's"})

    fitch = rtg.loc[rtg['agencia_tipo_rtg'].isin(['RTG_FITCH_NATIONAL_LT', 'RTG_FITCH_NATIONAL_SR_UNSECURED'])]
    fitch = pd.merge(fitch, depara, left_on='rtg', right_on='rtg', how='left')
    fitch = fitch.sort(['cnpj', 'cod_rtg'], ascending=[True, False])
    fitch = fitch.drop_duplicates(subset=['cnpj'], take_last=False)
    fitch = fitch.rename(columns={'rtg': "Fitch"})

    del fitch['cod_rtg']
    del moodys['agencia_tipo_rtg']
    del sp['agencia_tipo_rtg']
    del fitch['agencia_tipo_rtg']

    contraparte = pd.merge(contraparte, moodys, left_on='cnpj', right_on='cnpj', how='left')
    contraparte = pd.merge(contraparte, sp, left_on='cnpj', right_on='cnpj', how='left')
    contraparte = pd.merge(contraparte, fitch, left_on='cnpj', right_on='cnpj', how='left')
    contraparte = contraparte.reindex(columns=['cnpj', 'contraparte', 'mtm_info', "Moody's", "Standard & Poor's", "Fitch"])
    contraparte = contraparte.sort_values(by=['contraparte'], ascending=True)
    contraparte = contraparte.rename(columns={'mtm_info': 'Exposição (R$)'})

    if len(contraparte) != 0:
        contraparte.to_excel(writer, index=False, sheet_name='Rating por Contraparte', startrow=3, startcol=1)
        worksheet = writer.sheets['Rating por Contraparte']
        worksheet.set_column('D:D', 12, numero_float)
        contador_contraparte = 1

    # rating por isin
    base_isin = credito[['mtm_info', 'produto', 'isin', 'cnpj', 'contraparte', 'ativo']].copy()
    base_isin = base_isin.reset_index(level=None, drop=False, inplace=False, col_level=0, col_fill='')
    del base_isin['index']
    base_isin1 = pd.DataFrame(
        base_isin.groupby(['produto', 'isin', 'cnpj', 'contraparte', 'ativo'], as_index=False).sum())

    query = 'select distinct a.isin, a.agencia_tipo_rtg, a.rtg from projeto_inv.rating_isin as a right join (select max(data_bd) as data_bd from projeto_inv.rating_isin where dt_ref= "' + dt_base + '" ) as b on a.data_bd=b.data_bd;'
    rtg_isin = pd.read_sql(query, con=connection)

    moodys1 = rtg_isin.loc[
        rtg_isin['agencia_tipo_rtg'].isin(['RTG_MDY_NSR', 'RTG_MDY_NSR_SR_UNSECURED', 'RTG_MDY_NSR_SUBORDINATED'])]
    moodys1 = pd.merge(moodys1, depara, left_on='rtg', right_on='rtg', how='left')
    moodys1 = moodys1.sort(['isin', 'cod_rtg'], ascending=[True, False])
    moodys1 = moodys1.drop_duplicates(subset=['isin'], take_last=False)
    del moodys1['cod_rtg']
    moodys1 = moodys1.rename(columns={'rtg': "Moody's - Operação"})

    sp1 = rtg_isin[rtg_isin['agencia_tipo_rtg'] == 'RTG_SP_NATIONAL']
    sp1 = sp1.rename(columns={'rtg': "Standard & Poor's - Operação"})

    fitch1 = rtg_isin.loc[rtg_isin['agencia_tipo_rtg'].isin(
        ['RTG_FITCH_NATIONAL_LT', 'RTG_FITCH_NATIONAL', 'RTG_FITCH_NATIONAL_SR_UNSECURED',
         'RTG_FITCH_NATL_SUBORDINATED'])]
    fitch1 = pd.merge(fitch1, depara, left_on='rtg', right_on='rtg', how='left')
    fitch1 = fitch1.sort(['isin', 'cod_rtg'], ascending=[True, False])
    fitch1 = fitch1.drop_duplicates(subset=['isin'], take_last=False)
    del fitch1['cod_rtg']
    fitch1 = fitch1.rename(columns={'rtg': "Fitch's - Operação"})

    del moodys1['agencia_tipo_rtg']
    del sp1['agencia_tipo_rtg']
    del fitch1['agencia_tipo_rtg']

    base_isin1 = pd.merge(base_isin1, moodys1, left_on='isin', right_on='isin', how='left')
    base_isin1 = pd.merge(base_isin1, sp1, left_on='isin', right_on='isin', how='left')
    base_isin1 = pd.merge(base_isin1, fitch1, left_on='isin', right_on='isin', how='left')
    base_isin1 = pd.merge(base_isin1, moodys, left_on='cnpj', right_on='cnpj', how='left')
    base_isin1 = pd.merge(base_isin1, sp, left_on='cnpj', right_on='cnpj', how='left')
    base_isin1 = pd.merge(base_isin1, fitch, left_on='cnpj', right_on='cnpj', how='left')

    base_isin1 = base_isin1.reindex(
        columns=['cnpj', 'contraparte', 'isin', 'ativo', 'produto', 'mtm_info', "Moody's - Operação",
                 "Standard & Poor's - Operação", "Fitch's - Operação", "Moody's", "Standard & Poor's", "Fitch"])

    base_isin1 = base_isin1.rename(columns={'mtm_info': 'Exposição (R$)', "Moody's": "Moody's - Contraparte",
                                            "Standard & Poor's": "Standard & Poor's - Contraparte",
                                            "Fitch": "Fitch - Contraparte"})

    base_isin1 = base_isin1.sort_values(by=['produto', 'contraparte'], ascending=[True, True])

    if len(base_isin1) != 0:
        base_isin1.to_excel(writer, index=False, sheet_name='Rating por Isin', startrow=3, startcol=1)
        worksheet = writer.sheets['Rating por Isin']
        worksheet.set_column('G:G', 12, numero_float)
        contador_isin = 1

    # Relatório - QUADRO GERAL
    # Participação por classe de ativos
    # RELATÓRIO COM CURVA E MERCADO - tipo relatório = 'G'
    # Informações do quadro 419

    query = 'select a.* from projeto_inv.quaid_419 a right join (select id_relatorio_qo, tipo_relatorio, max(data_bd) as data_bd from projeto_inv.quaid_419 where id_relatorio_qo=' + id_relatorio_qo + ' and tipo_relatorio="G" group by 1,2) b on a.id_relatorio_qo=b.id_relatorio_qo and a.tipo_relatorio=b.tipo_relatorio and a.data_bd=b.data_bd;'
    quaid_419 = pd.read_sql(query, con=connection)
    quaid_419 = quaid_419[
        # Se tem linha duplicada, mais de um fator de risco, pega apenas a referente ao indexador principal
        quaid_419.EMFMULTIPLOFATOR == 0]
    id_quaid_419_df = quaid_419[['id_relatorio_quaid419']]
    id_relatorio_quaid419 = id_quaid_419_df.get_value(0, 'id_relatorio_quaid419').astype(str)

    # Seleção de indexadors que vem do quadro 419
    fator_419 = quaid_419[['EMFCODISIN', 'tipo_produto', 'indexador']].copy()
    fator_419 = fator_419.rename(columns={'indexador': 'indexador_q419'})
    del quaid_419['indexador']

    # Seleção dos indexadores que vem do quadro de operações
    qo_indexador = qo[['isin', 'produto', 'indexador']].copy()
    qo_indexador = qo_indexador[qo_indexador['isin'].notnull()].copy()
    qo_indexador['FTRCODIGO'] = qo_indexador['indexador']
    qo_indexador['FTRCODIGO'] = np.where(qo_indexador['produto'].isin(['Futuro']), 'PRE', qo_indexador['FTRCODIGO'])
    qo_indexador['FTRCODIGO'] = np.where(qo_indexador['produto'].str.contains('Termo'), 'PRE',
                                         qo_indexador['FTRCODIGO'])
    qo_indexador['FTRCODIGO'] = np.where(qo_indexador['FTRCODIGO'].isnull(), qo_indexador['produto'],
                                         qo_indexador['FTRCODIGO'])
    qo_indexador['FTRCODIGO'] = np.where(qo_indexador['FTRCODIGO'].isin(['DI', 'DI1', 'ANB', 'ANBID', 'Anbid']), 'CDI',
                                         qo_indexador['FTRCODIGO'])
    qo_indexador['FTRCODIGO'] = np.where(qo_indexador['FTRCODIGO'].isin(['IAP', 'IPC']), 'IPCA',
                                         qo_indexador['FTRCODIGO'])
    qo_indexador['FTRCODIGO'] = np.where(qo_indexador['FTRCODIGO'].isin(['IGM', 'IGP']), 'IGPM',
                                         qo_indexador['FTRCODIGO'])
    qo_indexador['FTRCODIGO'] = np.where(qo_indexador['FTRCODIGO'].isin(['SEL']), 'Selic', qo_indexador['FTRCODIGO'])
    qo_indexador['FTRCODIGO'] = np.where(qo_indexador['FTRCODIGO'].isin(['ajuste de futuro']), 'PRE',
                                         qo_indexador['FTRCODIGO'])
    qo_indexador['FTRCODIGO'] = np.where(qo_indexador['FTRCODIGO'].isin(['fundo']), 'Cotas de fundos',
                                         qo_indexador['FTRCODIGO'])
    qo_indexador['FTRCODIGO'] = np.where(qo_indexador['FTRCODIGO'].isin(['ações']), 'Bolsa', qo_indexador['FTRCODIGO'])

    qo_indexador = qo_indexador.drop_duplicates(subset=['isin', 'produto', 'indexador'], take_last=True)
    qo_indexador = qo_indexador.rename(columns={'isin': 'EMFCODISIN'})

    qo = pd.merge(qo, qo_indexador, left_on=['isin', 'produto', 'indexador'],
                  right_on=['EMFCODISIN', 'produto', 'indexador'], how='left')

    # RELATÓRIO SÓ A MERCADO - cálculo de duration
    query = 'select a.* from projeto_inv.quaid_419 a right join (select id_relatorio_qo, tipo_relatorio, max(data_bd) as data_bd from projeto_inv.quaid_419 where id_relatorio_qo=' + id_relatorio_qo + ' and tipo_relatorio="R" group by 1,2) b on a.id_relatorio_qo=b.id_relatorio_qo and a.tipo_relatorio=b.tipo_relatorio and a.data_bd=b.data_bd;'
    quaid_419_r = pd.read_sql(query, con=connection)
    quaid_419_r['expo'] = np.where(quaid_419_r['TPFOPERADOR'] == '-', -1 * quaid_419_r['EMFVLREXPRISCO'],
                                   quaid_419_r['EMFVLREXPRISCO'])
    quaid_419_r['prazo_ponderado'] = quaid_419_r['expo'] * quaid_419_r['EMFPRAZOFLUXO']
    duration = quaid_419_r[['FTRCODIGO', 'expo', 'prazo_ponderado', 'tipo_produto']].copy()
    duration['FTRCODIGO'] = np.where(duration['FTRCODIGO'].isnull(), duration['tipo_produto'], duration['FTRCODIGO'])
    duration['FTRCODIGO'] = np.where(duration['FTRCODIGO'].isin(['TD1', 'TD2']), 'CDI', duration['FTRCODIGO'])
    duration['FTRCODIGO'] = np.where(duration['FTRCODIGO'].isin(['JI1']), 'IPCA', duration['FTRCODIGO'])
    duration['FTRCODIGO'] = np.where(duration['FTRCODIGO'].isin(['JI2']), 'IGPM', duration['FTRCODIGO'])
    duration['FTRCODIGO'] = np.where(duration['FTRCODIGO'].isin(['TS1']), 'Selic', duration['FTRCODIGO'])
    duration['FTRCODIGO'] = np.where(duration['FTRCODIGO'].isin(['JJ1']), 'PRE', duration['FTRCODIGO'])
    duration['FTRCODIGO'] = np.where(duration['FTRCODIGO'].isin(['FF1']), 'Cotas de fundos', duration['FTRCODIGO'])
    duration['FTRCODIGO'] = np.where(duration['FTRCODIGO'].isin(['AA1']), 'Bolsa', duration['FTRCODIGO'])

    duration_ativa = duration[duration.expo >= 0].copy()
    duration_ativa_fator = duration_ativa.groupby('FTRCODIGO', as_index=False).sum()
    duration_ativa_fator['duration'] = duration_ativa_fator['prazo_ponderado'] / duration_ativa_fator['expo']
    duration_passiva = duration[duration.expo < 0].copy()
    duration_passiva_fator = duration_passiva.groupby('FTRCODIGO', as_index=False).sum()
    duration_passiva_fator['duration'] = duration_passiva_fator['prazo_ponderado'] / duration_passiva_fator['expo']

    # DV100
    query = 'select distinct a.codigo_isin, a.data_mtm, a.DV100,a.mtm, a.data_bd from projeto_inv.mtm_renda_fixa a right join (select distinct codigo_isin, data_mtm, max(data_bd) as data_bd from projeto_inv.mtm_renda_fixa where data_mtm="' + dt_base + '" group by 1,2 ) b on a.codigo_isin=b.codigo_isin and a.data_bd=b.data_bd and a.data_mtm=b.data_mtm;'
    mtm_rf = pd.read_sql(query, con=connection)
    del mtm_rf['data_bd']
    del mtm_rf['data_mtm']
    mtm_rf['perc_dv100'] = mtm_rf.DV100 / mtm_rf.mtm

    del mtm_rf['mtm']
    qo = pd.merge(qo, mtm_rf, left_on='isin', right_on='codigo_isin', how='left')
    qo['dv100'] = np.where((qo['caracteristica'] == 'N') & (qo['perc_dv100'] > 0), qo['perc_dv100'] * qo['mtm_info'], 0)
    del qo['codigo_isin']

    # fidc
    fundos = qo.loc[qo['produto'].isin(['fundo'])].copy()
    fundos['fundo_final'] = np.where(fundos['fundo_ult_nivel'].isnull(), fundos['fundo'], fundos['fundo_ult_nivel'])
    fidc = fundos[fundos['fundo_final'].str.contains('FIDC|DIREITOS|CREDITÓRIO|CREDITORIOS|DIREITOS')]
    fidc1 = pd.DataFrame()
    fidc1['isin'] = fidc['isin'].unique()
    fidc1['flag_fidc'] = 1

    qo = pd.merge(qo, fidc1, how='left', left_on='isin', right_on='isin')
    qo['dv100'] = np.where((qo['flag_fidc'] == 1) & (qo['DV100'] > 0), qo['DV100'] * qo['mtm_info'], qo['dv100'])
    qo['dv100'] = qo['dv100'].fillna(0)

    del qo['DV100']
    del qo['flag_fidc']

    expo = qo[['FTRCODIGO', 'mtm_info', 'dv100']]

    expo1 = expo.groupby(["FTRCODIGO"], as_index=False).sum()
    expo2 = expo1[(expo1.FTRCODIGO != 'IMO') & (expo1.FTRCODIGO != 'DPV') & (expo1.FTRCODIGO != 'PSR')]
    expo2 = pd.merge(expo2, duration_ativa_fator, left_on='FTRCODIGO', right_on=['FTRCODIGO'], how='left')
    expo2 = expo2.rename(columns={'duration': 'duration_ativa'})
    expo2 = pd.merge(expo2, duration_passiva_fator, left_on='FTRCODIGO', right_on=['FTRCODIGO'], how='left')
    expo2 = expo2.rename(columns={'duration': 'duration_passiva'})
    expo2['duration_ativa'] = expo2['duration_ativa'].fillna(0)
    expo2['duration_passiva'] = expo2['duration_passiva'].fillna(0)

    expo3 = expo2.copy()
    total_expo = sum(expo3['mtm_info'])
    expo3 = expo3.reindex(columns=['FTRCODIGO', 'mtm_info', 'dv100', 'duration_ativa', 'duration_passiva'])
    expo3['FTRCODIGO'] = expo3.FTRCODIGO.fillna('sem fator de risco')
    expo3 = expo3.groupby('FTRCODIGO', as_index=False).sum()
    expo3['DV100 (%)'] = expo3['dv100'] / expo3['mtm_info']
    del expo3['dv100']
    expo3 = expo3.rename(
        columns={'FTRCODIGO': 'Fator de Risco', 'mtm_info': 'MtM (R$)', 'duration_ativa': 'Duration Ativa (DU)',
                 'duration_passiva': 'Duration Passiva (DU)'})
    expo3['Participação (%)'] = expo3['MtM (R$)'] / total_expo

    if len(expo3) != 0:
        expo3.to_excel(writer, index=False, sheet_name='Quadro Geral', startrow=3, startcol=1)

    # NORMAL
    query = 'select a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, tipo_segmento, norm_stress, max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + str(
        id_relatorio_quaid419) + ' and tipo_var="Componente" and tipo_alocacao="Total" and tipo_segmento="Total" and norm_stress="normal" group by 1,2,3,4,5,6) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.vertice=b.vertice and a.tipo_alocacao=b.tipo_alocacao and a.tipo_segmento=b.tipo_segmento and a.norm_stress = b.norm_stress and a.data_bd=b.data_bd ; '
    varcomponente = pd.read_sql(query, con=connection)

    varcomponente['fator_risco'] = varcomponente['vertice'].str.split("_")
    varcomponente['vertice1'] = varcomponente['fator_risco'].str[1]
    varcomponente['fator_risco'] = varcomponente['fator_risco'].str[0]

    varcomponente_tab = varcomponente[['fator_risco', 'vertice1', 'var']]
    varcomponente_tab['vertice1'] = varcomponente_tab['vertice1'].fillna(1)
    varcomponente_tab = varcomponente_tab.groupby(['fator_risco', 'vertice1'], as_index=False).sum()
    varcomponente_tab = varcomponente_tab.rename(columns={'var': 'VaR Componente (R$)'})
    varcomponente_normal = varcomponente_tab.copy()

    # ESTRESSADO
    query = 'select a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, tipo_segmento, norm_stress, max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + str(
        id_relatorio_quaid419) + ' and tipo_var="Componente" and tipo_alocacao="Total" and tipo_segmento="Total" and norm_stress="estressado" group by 1,2,3,4,5,6) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.vertice=b.vertice and a.tipo_alocacao=b.tipo_alocacao and a.tipo_segmento=b.tipo_segmento and a.norm_stress = b.norm_stress and a.data_bd=b.data_bd ; '
    varcomponente = pd.read_sql(query, con=connection)
    varcomponente['fator_risco'] = varcomponente['vertice'].str.split("_")
    varcomponente['vertice1'] = varcomponente['fator_risco'].str[1]
    varcomponente['fator_risco'] = varcomponente['fator_risco'].str[0]

    varcomponente_tab = varcomponente[['fator_risco', 'vertice1', 'var']]
    varcomponente_tab = varcomponente_tab[varcomponente_tab['var'] != 0]
    varcomponente_tab['vertice1'] = varcomponente_tab['vertice1'].fillna(1)
    varcomponente_tab = varcomponente_tab.groupby(['fator_risco', 'vertice1'], as_index=False).sum()
    varcomponente_tab = varcomponente_tab.rename(columns={'var': 'VaR Componente Estressado (R$)'})
    varcomponente_estressado = varcomponente_tab.copy()

    varcomponente_tab = pd.merge(varcomponente_normal, varcomponente_estressado, left_on=['fator_risco', 'vertice1'],
                                 right_on=['fator_risco', 'vertice1'], how='left')

    # Títulos
    n_linhas_expo = len(expo3)

    del qo['EMFCODISIN']

    titulos = pd.DataFrame()

    ano = int(dt_base[0:4])
    mes = int(dt_base[5:7])
    dia = int(dt_base[8:10])
    qo['delta'] = pd.to_datetime(qo['dt_vencto']) - datetime.datetime(ano, mes, dia)
    qo['delta'] = qo['delta'].dt.days

    qo_venctos = qo[qo['produto'].isin(['titulo público', 'título privado', 'debênture'])].copy()

    # NTN-B curtas (até 5 anos)
    ntnb_curta = qo_venctos[
        (qo_venctos.produto == 'titulo público') & (qo_venctos.delta <= 360 * 5) & (qo_venctos.FTRCODIGO == 'IPCA')]
    ntnb_curta = ntnb_curta[['header_id', 'mtm_info', 'quantidade']].copy()
    ntnb_curta_fim = ntnb_curta.groupby(['header_id'], as_index=False).sum()
    ntnb_curta_fim['titulos'] = 'NTNBs Curtas (até 5 anos)'
    if len(ntnb_curta_fim) > 0:
        titulos = titulos.append(ntnb_curta_fim)

    # NTN-B longa (acima 5 anos)
    ntnb_longa = qo_venctos[
        (qo_venctos.produto == 'titulo público') & (qo_venctos.delta > 360 * 5) & (qo_venctos.FTRCODIGO == 'IPCA')]
    ntnb_longa = ntnb_longa[['header_id', 'mtm_info', 'quantidade']].copy()
    ntnb_longa_fim = ntnb_longa.groupby(['header_id'], as_index=False).sum()
    ntnb_longa_fim['titulos'] = 'NTNBs Longas (acima 5 anos)'
    if len(ntnb_longa_fim) > 0:
        titulos = titulos.append(ntnb_longa_fim)

    # NTN-C curtas (até 5 anos)
    ntnc_curta = qo_venctos[
        (qo_venctos.produto == 'titulo público') & (qo_venctos.delta <= 360 * 5) & (qo_venctos.FTRCODIGO == 'IGPM')]
    ntnc_curta = ntnc_curta[['header_id', 'mtm_info', 'quantidade']].copy()
    ntnc_curta_fim = ntnc_curta.groupby(['header_id'], as_index=False).sum()
    ntnc_curta_fim['titulos'] = 'NTNCs Curtas (até 5 anos)'
    if len(ntnc_curta_fim) > 0:
        titulos = titulos.append(ntnc_curta_fim)

    # NTN-C longa (acima 5 anos)
    ntnc_longa = qo_venctos[
        (qo_venctos.produto == 'titulo público') & (qo_venctos.delta <= 360 * 5) & (qo_venctos.FTRCODIGO == 'IGPM')]
    ntnc_longa = ntnc_longa[['header_id', 'mtm_info', 'quantidade']].copy()
    ntnc_longa_fim = ntnc_longa.groupby(['header_id'], as_index=False).sum()
    ntnc_longa_fim['titulos'] = 'NTNCs Longas (acima 5 anos)'
    if len(ntnc_longa_fim) > 0:
        titulos = titulos.append(ntnc_longa_fim)

    # crédito privado indexado ao CDI
    credito = qo_venctos[(qo_venctos.produto == 'debênture') | (qo_venctos.produto == 'título privado')]
    credito_cdi = credito[(credito.FTRCODIGO == 'CDI')]
    credito_cdi = credito_cdi[['header_id', 'mtm_info', 'quantidade']].copy()
    credito_cdi_fim = credito_cdi.groupby(['header_id'], as_index=False).sum()
    credito_cdi_fim['titulos'] = 'Crédito privado indexado ao CDI'
    if len(credito_cdi_fim) > 0:
        titulos = titulos.append(credito_cdi_fim)

    # crédito privado indexado ao IPCA
    credito_ipca = credito[credito.FTRCODIGO == 'IPCA']
    credito_ipca = credito_ipca[['header_id', 'mtm_info', 'quantidade']].copy()
    credito_ipca_fim = credito_ipca.groupby(['header_id'], as_index=False).sum()
    credito_ipca_fim['titulos'] = 'Crédito privado indexado ao IPCA'
    if len(credito_ipca_fim) > 0:
        titulos = titulos.append(credito_ipca_fim)

    # crédito privado indexado ao IGPM
    credito_igpm = credito[credito.FTRCODIGO == 'IGPM']
    credito_igpm = credito_igpm[['header_id', 'mtm_info', 'quantidade']].copy()
    credito_igpm_fim = credito_igpm.groupby(['header_id'], as_index=False).sum()
    credito_igpm_fim['titulos'] = 'Crédito privado indexado ao IGP-M'
    if len(credito_igpm_fim) > 0:
        titulos = titulos.append(credito_igpm_fim)

    if len(titulos) > 0:
        tot_titulos = sum(titulos.mtm_info)
        titulos['perc'] = titulos['mtm_info'] / tot_titulos
        titulos = titulos.reindex(columns=['titulos', 'mtm_info', 'perc', 'quantidade'])
        titulos = titulos.rename(columns={'titulos': 'Títulos', 'mtm_info': 'MtM (R$)', 'perc': 'Participação (%)',
                                          'quantidade': 'Quantidade'})
        titulos.to_excel(writer, index=False, sheet_name='Quadro Geral', startrow=3 + n_linhas_expo + 4, startcol=1)
        contador_titulos = 1

        n_cat_prod = 3 + n_linhas_expo + 4 + len(titulos) + 1

    else:
        n_cat_prod = 3 + n_linhas_expo + 4 + 1

    del qo_venctos

    # Grupo de produtos
    qo['cat_produto'] = qo['produto']
    qo['cat_produto'] = np.where(qo.produto.str.contains('despesa'), 'despesa', qo['cat_produto'])
    qo['cat_produto'] = np.where(qo.produto.str.contains('Termo'), 'termo', qo['cat_produto'])
    qo['cat_produto'] = np.where(qo.produto.str.contains('Opções'), 'opções', qo['cat_produto'])

    grupo_prod = qo[['cat_produto', 'mtm_info']]
    grupo_prod = grupo_prod.groupby('cat_produto', as_index=False).sum()
    tot_grupo_prod = sum(grupo_prod.mtm_info)

    grupo_prod['mtm_perc'] = grupo_prod['mtm_info'] / tot_grupo_prod
    grupo_prod = grupo_prod.rename(
        columns={'mtm_info': 'MtM (R$)', 'mtm_perc': 'MtM (%)', 'cat_produto': 'Tipo de produto'})

    if len(grupo_prod) != 0:
        grupo_prod.to_excel(writer, index=False, sheet_name='Quadro Geral', startrow=n_cat_prod + 3, startcol=1)
        contador_grupo_prod = 1

    # ações - níveis diferenciados de governança
    n_acoes = n_cat_prod + len(grupo_prod) + 1
    query = 'select distinct  a.codigo_emissor, a.segmento, a.data_bd from projeto_inv.bmf_empresas_listadas a right join (select distinct codigo_emissor, max(data_bd) as data_bd from projeto_inv.bmf_empresas_listadas group by 1) b on a.codigo_emissor=b.codigo_emissor and a.data_bd=b.data_bd;'
    governanca = pd.read_sql(query, con=connection)

    acoes = qo[qo.produto == 'ações']
    acoes['cod_emissor'] = acoes['ativo'].str[0:4]
    acoes1 = acoes[['mtm_info', 'quantidade', 'cod_emissor']]
    acoes2 = acoes1.groupby('cod_emissor', as_index=False).sum()
    acoes2 = pd.merge(acoes2, governanca, left_on='cod_emissor', right_on='codigo_emissor', how='left')

    acoes3 = acoes2[['segmento', 'quantidade', 'mtm_info']]
    acoes3['segmento'] = acoes3['segmento'].fillna('SEM')
    acoes4 = acoes3.groupby('segmento', as_index=False).sum()

    total_acoes = sum(acoes4['mtm_info'])
    acoes4['partip'] = acoes4['mtm_info'] / total_acoes

    acoes4.segmento = acoes4.segmento.replace('NM', 'Novo Mercado')
    acoes4.segmento = acoes4.segmento.replace('N1', 'Nível 1 de Governança Corporativa')
    acoes4.segmento = acoes4.segmento.replace('N2', 'Nível 2 de Governança Corporativa')
    acoes4.segmento = acoes4.segmento.replace('MA', 'Bovespa Mais')
    acoes4.segmento = acoes4.segmento.replace('M2', 'Bovespa Mais Nível 2')
    acoes4.segmento = acoes4.segmento.replace('MB', 'Cia. Balcão Org. Tradicional')
    acoes4.segmento = acoes4.segmento.replace('DR1', 'BDR Nível 1')
    acoes4.segmento = acoes4.segmento.replace('DR2', 'BDR Nível 2')
    acoes4.segmento = acoes4.segmento.replace('DR3', 'BDR Nível 3')
    acoes4.segmento = acoes4.segmento.replace('DRN', 'BDR Não Patrocinado')
    acoes4.segmento = acoes4.segmento.replace('SEM', 'Não listados nos níveis diferenciados')
    acoes4 = acoes4.reindex(columns=['segmento', 'mtm_info', 'partip', 'quantidade'])
    acoes4 = acoes4.rename(
        columns={'segmento': 'Níveis diferenciados de governança', 'mtm_info': 'Valor de Mercado (R$)',
                 'partip': 'Participação (%)', 'quantidade': 'Quantidade'})

    if len(acoes4) != 0:
        acoes4.to_excel(writer, index=False, sheet_name='Quadro Geral', startrow=n_acoes + 6, startcol=1)
        contador_acoes = 1

    worksheet = writer.sheets['Quadro Geral']
    worksheet.set_column('C:C', 12, numero_float)
    worksheet.set_column('E:E', 12, numero_float)
    worksheet.set_column('D13:D100000', 12, percent)
    worksheet.set_column('D1:D12', 12, numero_float)
    worksheet.set_column('F:G', 12, percent)

    # Relatório - Descasamento -> baseado no quadro gerencial
    query = 'select a.* from projeto_inv.vetor_exposicoes a right join (select id_relatorio_quaid419, max(data_bd) as data_bd from projeto_inv.vetor_exposicoes where id_relatorio_quaid419=' + id_relatorio_quaid419 + ' group by 1) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.data_bd=b.data_bd;'
    vetor_expo = pd.read_sql(query, con=connection)
    vetor_expo1 = vetor_expo[(vetor_expo['valor_exposicao'] != 0) & (~vetor_expo['valor_exposicao'].isnull())].copy()
    vetor_expo1 = vetor_expo1.reset_index(level=None, drop=False, inplace=False, col_level=0, col_fill='')
    del vetor_expo1['index']
    vetor_expo2 = pd.DataFrame(vetor_expo1.vertice.str.split('_', 1).tolist(), columns=['fator_risco', 'num_vertice'])
    vetor_expo3 = vetor_expo1.join(vetor_expo2)

    lista_fatores_risco = pd.DataFrame(pd.Series(vetor_expo3.fator_risco.ravel()).unique())
    lista_fatores_risco = lista_fatores_risco.rename(columns={0: 'fator'})

    lista_fatores_risco1 = lista_fatores_risco.loc[lista_fatores_risco['fator'].isin(['PRE', 'IPCA', 'IGPM'])]
    lista_fatores_risco1 = lista_fatores_risco1.reset_index(level=None, drop=False, inplace=False, col_level=0, col_fill='')

    for i in range(0, len(lista_fatores_risco1)):
        nome_fator = lista_fatores_risco1['fator'][i]
        fluxo_caixa_expo = vetor_expo3.loc[vetor_expo3['fator_risco'].isin([nome_fator])]
        fluxo_caixa_expo = fluxo_caixa_expo[['fator_risco', 'num_vertice', 'valor_exposicao']]
        fluxo_caixa_expo['num_vertice'] = fluxo_caixa_expo['num_vertice'].astype(int)
        fluxo_caixa_expo = fluxo_caixa_expo.sort_values(by=['num_vertice'], ascending=True)
        fluxo_caixa_expo = fluxo_caixa_expo.rename(
            columns={'fator_risco': 'Fator de Risco', 'num_vertice': 'Vértice', 'valor_exposicao': 'Exposição (R$)'})
        fluxo_caixa_expo.to_excel(writer, index=False, sheet_name='descasamento_' + nome_fator, startrow=3, startcol=1)

        worksheet = writer.sheets['descasamento_' + nome_fator]
        worksheet.set_column('D:D', 12, numero_float)
        contador_descasamento = contador_descasamento + nome_fator

    # Relatório - Resumo
    # Infos quadro de operações
    pl_qo = qo[qo.produto != 'Futuro']
    pl_info = sum(pl_qo['mtm_info'])
    pl_calc = sum(pl_qo['mtm_regra_xml'])

    query = 'select a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, tipo_segmento, norm_stress, max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + str(
        id_relatorio_quaid419) + ' and tipo_var="Total" and vertice="Total" and tipo_alocacao="Total" and tipo_segmento="Total" and norm_stress = "normal" group by 1,2,3,4,5,6) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.tipo_alocacao=b.tipo_alocacao and a.tipo_segmento=b.tipo_segmento and a.norm_stress = b.norm_stress and a.data_bd=b.data_bd; '
    vargeral_normal = pd.read_sql(query, con=connection)

    query = 'select a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, tipo_segmento, norm_stress, max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + str(
        id_relatorio_quaid419) + ' and tipo_var="Total" and vertice="Total" and tipo_alocacao="Total" and tipo_segmento="Total" and norm_stress = "estressado" group by 1,2,3,4,5,6) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.tipo_alocacao=b.tipo_alocacao and a.tipo_segmento=b.tipo_segmento and a.norm_stress = b.norm_stress and a.data_bd=b.data_bd; '
    vargeral_estressado = pd.read_sql(query, con=connection)

    var_par = vargeral_normal[['nivel_confianca', 'horizonte_tempo', 'var']]
    var_par['modelo'] = 'Paramétrico'
    var_par['nivel_confianca'] = '99%'

    var_par = var_par.rename(columns={'modelo': 'Modelo', 'nivel_confianca': 'Nível Confiança (%)',
                                      'horizonte_tempo': 'Horizonte de Tempo (DU)'})
    var_par = var_par.reindex(columns=['Modelo', 'Nível Confiança (%)', 'Horizonte de Tempo (DU)'])
    var_par.to_excel(writer, index=False, sheet_name='Resumo - Visão Consolidada', startrow=3, startcol=1)

    var_resumo = pd.DataFrame(columns=['Patrimônio Líquido (R$)', 'Value at Risk (R$)', 'Value at Risk (%)',
                                       'Value at Risk - Estressado (R$)', 'Value at Risk - Estressado (%)'])
    var_resumo['Value at Risk - Estressado (R$)'] = vargeral_estressado['var']
    var_resumo['Value at Risk - Estressado (%)'] = vargeral_estressado['var'] / pl_info
    var_resumo['Patrimônio Líquido (R$)'] = pl_info
    var_resumo['Value at Risk (R$)'] = vargeral_normal['var']
    var_resumo['Value at Risk (%)'] = vargeral_normal['var'] / pl_info

    var_resumo = var_resumo.reindex(columns=['Patrimônio Líquido (R$)', 'Value at Risk (R$)', 'Value at Risk (%)',
                                             'Value at Risk - Estressado (R$)', 'Value at Risk - Estressado (%)'])
    var_resumo.to_excel(writer, index=False, sheet_name='Resumo - Visão Consolidada', startrow=8, startcol=1)

    worksheet = writer.sheets['Resumo - Visão Consolidada']
    worksheet.set_column('B:B', 12, numero_float)
    worksheet.set_column('E:E', 12, numero_float)

    # Var segmento NORMAL
    query = 'select a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, norm_stress, max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + str(
        id_relatorio_quaid419) + ' and tipo_var="Total" and vertice="Total" and tipo_alocacao="segmento" and norm_stress="normal" group by 1,2,3,4,5) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.tipo_alocacao=b.tipo_alocacao and a.norm_stress=b.norm_stress and a.data_bd=b.data_bd; '
    var_segmento = pd.read_sql(query, con=connection)
    var_seg = var_segmento[['tipo_segmento', 'var']]
    var_seg['perc_var'] = var_seg['var'] / pl_info
    del var_seg['var']
    var_seg = var_seg.rename(columns={'tipo_segmento': 'Tipo de Segmento', 'perc_var': '%VaR'})
    var_seg_norm = var_seg.copy()

    # Var segmento ESTRESSADO
    query = 'select a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, norm_stress, max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + str(
        id_relatorio_quaid419) + ' and tipo_var="Total" and vertice="Total" and tipo_alocacao="segmento" and norm_stress="estressado" group by 1,2,3,4,5) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.tipo_alocacao=b.tipo_alocacao and a.norm_stress=b.norm_stress and a.data_bd=b.data_bd; '
    var_segmento = pd.read_sql(query, con=connection)
    var_seg = var_segmento[['tipo_segmento', 'var']]
    var_seg['perc_var'] = var_seg['var'] / pl_info
    del var_seg['var']
    var_seg = var_seg.rename(columns={'tipo_segmento': 'Tipo de Segmento', 'perc_var': '%VaR Estressado'})
    var_seg_stress = var_seg.copy()

    var_seg = pd.merge(var_seg_norm, var_seg_stress, left_on=['Tipo de Segmento'], right_on=['Tipo de Segmento'],
                       how='left')

    var_seg.to_excel(writer, index=False, sheet_name='Resumo - Visão Consolidada', startrow=13, startcol=1)

    worksheet.set_column('C13:D18', 12, percent)

    # Var marginal
    # NORMAL
    query = 'select a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, tipo_segmento, norm_stress, max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + id_relatorio_quaid419 + ' and tipo_var="Marginal" and tipo_alocacao="Total" and tipo_segmento="Total" and norm_stress="normal" group by 1,2,3,4,5,6) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.vertice=b.vertice and a.tipo_alocacao=b.tipo_alocacao and a.tipo_segmento=b.tipo_segmento and a.norm_stress=b.norm_stress and a.data_bd=b.data_bd ; '
    varmarginal_norm = pd.read_sql(query, con=connection)
    varmarginal_norm = varmarginal_norm[['vertice', 'var']].copy()
    varmarginal_norm = varmarginal_norm.rename(columns={'var': 'var_norm'})

    # ESTRESSADO
    query = 'select a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, tipo_segmento, norm_stress, max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + id_relatorio_quaid419 + ' and tipo_var="Marginal" and tipo_alocacao="Total" and tipo_segmento="Total" and norm_stress="estressado" group by 1,2,3,4,5,6) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.vertice=b.vertice and a.tipo_alocacao=b.tipo_alocacao and a.tipo_segmento=b.tipo_segmento and a.norm_stress=b.norm_stress and a.data_bd=b.data_bd ; '
    varmarginal_stress = pd.read_sql(query, con=connection)
    varmarginal_stress = varmarginal_stress[['vertice', 'var']].copy()
    varmarginal_stress = varmarginal_stress.rename(columns={'var': 'var_stress'})
    varmarginal = pd.merge(varmarginal_norm, varmarginal_stress, left_on=['vertice'], right_on=['vertice'], how='left')
    varmarginal['fator_risco'] = varmarginal['vertice'].str.split("_")
    varmarginal['vertice1'] = varmarginal['fator_risco'].str[1]
    varmarginal['fator_risco'] = varmarginal['fator_risco'].str[0]

    varmarginal_tab = varmarginal[['fator_risco', 'vertice1', 'var_norm', 'var_stress']]
    varmarginal_tab = varmarginal_tab[varmarginal_tab['var_norm'] != 0]
    varmarginal_tab['vertice1'] = varmarginal_tab['vertice1'].fillna(1)
    varmarginal_tab = varmarginal_tab.groupby(['fator_risco', 'vertice1'], as_index=False).sum()
    varmarginal_tab = varmarginal_tab.rename(
        columns={'var_norm': 'VaR Marginal (R$)', 'var_stress': 'VaR Marginal Estressado (R$)'})

    var_comp_marg = pd.merge(varcomponente_tab, varmarginal_tab, left_on=['fator_risco', 'vertice1'],
                             right_on=['fator_risco', 'vertice1'], how='left')
    var_comp_marg['Vértice'] = var_comp_marg['vertice1'].astype(int)
    del var_comp_marg['vertice1']
    var_comp_marg = var_comp_marg.sort(['fator_risco', 'Vértice'])
    var_comp_marg = var_comp_marg.rename(columns={'fator_risco': 'Fator de Risco'})
    var_comp_marg = var_comp_marg.reindex(
        columns=['Fator de Risco', 'Vértice', 'VaR Componente (R$)', 'VaR Componente Estressado (R$)',
                 'VaR Marginal (R$)', 'VaR Marginal Estressado (R$)'])
    var_comp_marg.to_excel(writer, index=False, sheet_name='Resumo - Visão Consolidada', startrow=22, startcol=1)

    worksheet = writer.sheets['Resumo - Visão Consolidada']
    worksheet.set_column('C:C', 12, numero_float)
    worksheet.set_column('E:E', 12, numero_float)
    worksheet.set_column('D:D', 12, numero_float)
    worksheet.set_column('C9:C10', 12, numero_float)

    # VaR por Produto
    # Var ESTRESSADO
    query = 'select a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, norm_stress, max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + str(
        id_relatorio_quaid419) + ' and tipo_var="Total" and vertice="Total" and tipo_alocacao="produto" and norm_stress="normal" group by 1,2,3,4,5) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.tipo_alocacao=b.tipo_alocacao and a.norm_stress=b.norm_stress and a.data_bd=b.data_bd; '
    var_prod_tot = pd.read_sql(query, con=connection)
    del var_prod_tot['id_var']
    var_prod_tot = var_prod_tot[['tipo_segmento', 'var']]
    var_prod_tot['% VaR'] = var_prod_tot['var'] / pl_info
    del var_prod_tot['var']
    var_prod_norm = var_prod_tot.copy()

    # Var NORMAL
    query = 'select a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, norm_stress, max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + str(
        id_relatorio_quaid419) + ' and tipo_var="Total" and vertice="Total" and tipo_alocacao="produto" and norm_stress="estressado" group by 1,2,3,4,5) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.tipo_alocacao=b.tipo_alocacao and a.norm_stress=b.norm_stress and a.data_bd=b.data_bd; '
    var_prod_tot = pd.read_sql(query, con=connection)
    del var_prod_tot['id_var']
    var_prod_tot = var_prod_tot[['tipo_segmento', 'var']]
    var_prod_tot['% VaR Estressado'] = var_prod_tot['var'] / pl_info
    del var_prod_tot['var']
    var_prod_stress = var_prod_tot.copy()

    var_prod_tot = pd.merge(var_prod_norm, var_prod_stress, left_on=['tipo_segmento'], right_on=['tipo_segmento'],
                            how='left')
    var_prod_tot = var_prod_tot.rename(columns={'tipo_segmento': 'Segmento'})
    var_prod_tot.to_excel(writer, index=False, sheet_name='VaR_segmento', startrow=3, startcol=1)

    # Var componente por produto
    # NORMAL
    query = 'select distinct a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, norm_stress,  max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + str(
        id_relatorio_quaid419) + ' and tipo_var="Componente" and vertice<>"Total" and tipo_alocacao="produto" and norm_stress="normal" group by 1,2,3,4,5) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.tipo_alocacao=b.tipo_alocacao and a.norm_stress=b.norm_stress and a.data_bd=b.data_bd;'
    var_prod_componente = pd.read_sql(query, con=connection)
    var_prod_componente = var_prod_componente[['tipo_segmento', 'vertice', 'var']]
    var_prod_componente = var_prod_componente.rename(columns={'var': 'VaR Componente (R$)'})
    var_prod_componente_normal = var_prod_componente.copy()

    # ESTRESSADO
    query = 'select distinct a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, norm_stress,  max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + str(
        id_relatorio_quaid419) + ' and tipo_var="Componente" and vertice<>"Total" and tipo_alocacao="produto" and norm_stress="estressado" group by 1,2,3,4,5) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.tipo_alocacao=b.tipo_alocacao and a.norm_stress=b.norm_stress and a.data_bd=b.data_bd;'
    var_prod_componente = pd.read_sql(query, con=connection)
    var_prod_componente = var_prod_componente[['tipo_segmento', 'vertice', 'var']]
    var_prod_componente = var_prod_componente.rename(columns={'var': 'VaR Componente Estressado (R$)'})
    var_prod_componente_estressado = var_prod_componente.copy()

    var_prod_componente = pd.merge(var_prod_componente_normal, var_prod_componente_estressado,
                                   left_on=['tipo_segmento', 'vertice'], right_on=['tipo_segmento', 'vertice'],
                                   how='left')

    # Var marginal por produto
    # NORMAL
    query = 'select distinct a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, norm_stress,  max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + str(
        id_relatorio_quaid419) + ' and tipo_var="Marginal" and vertice<>"Total" and tipo_alocacao="produto" and norm_stress="normal" group by 1,2,3,4,5) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.tipo_alocacao=b.tipo_alocacao and a.norm_stress=b.norm_stress and a.data_bd=b.data_bd;'
    var_prod_marginal = pd.read_sql(query, con=connection)
    var_prod_marginal = var_prod_marginal[['tipo_segmento', 'vertice', 'var']]
    var_prod_marginal = var_prod_marginal.rename(columns={'var': 'VaR Marginal (R$)'})
    var_prod_marginal_normal = var_prod_marginal.copy()

    # ESTRESSADO
    query = 'select distinct a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, norm_stress,  max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419=' + str(
        id_relatorio_quaid419) + ' and tipo_var="Marginal" and vertice<>"Total" and tipo_alocacao="produto" and norm_stress="estressado" group by 1,2,3,4,5) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.tipo_alocacao=b.tipo_alocacao and a.norm_stress=b.norm_stress and a.data_bd=b.data_bd;'
    var_prod_marginal = pd.read_sql(query, con=connection)
    var_prod_marginal = var_prod_marginal[['tipo_segmento', 'vertice', 'var']]
    var_prod_marginal = var_prod_marginal.rename(columns={'var': 'VaR Marginal Estressado (R$)'})
    var_prod_marginal_estressado = var_prod_marginal.copy()

    var_prod_marginal = pd.merge(var_prod_marginal_normal, var_prod_marginal_estressado,
                                 left_on=['tipo_segmento', 'vertice'], right_on=['tipo_segmento', 'vertice'],
                                 how='left')

    var_produto = pd.merge(var_prod_componente, var_prod_marginal, left_on=['tipo_segmento', 'vertice'],
                           right_on=['tipo_segmento', 'vertice'], how='left')
    var_produto['vertice'] = var_produto['vertice'].str.split('_')
    var_produto['fator_risco'] = var_produto['vertice'].str[0]
    var_produto['Vértice (DU)'] = var_produto['vertice'].str[1]
    var_produto['Vértice (DU)'] = var_produto['Vértice (DU)'].fillna('1')
    var_produto['Vértice (DU)'] = var_produto['Vértice (DU)'].astype(int)

    del var_produto['vertice']

    var_produto = var_produto.sort(['tipo_segmento', 'fator_risco', 'Vértice (DU)'])
    var_produto = var_produto.rename(columns={'fator_risco': 'Fator de Risco', 'tipo_segmento': 'Produto'})
    var_produto = var_produto.reindex(
        columns=['Produto', 'Fator de Risco', 'Vértice (DU)', 'VaR Componente (R$)', 'VaR Componente Estressado (R$)',
                 'VaR Marginal (R$)', 'VaR Marginal Estressado (R$)'])
    var_produto = var_produto[
        (var_produto['VaR Componente (R$)'] != 0) & (var_produto['VaR Componente (R$)'].notnull())].copy()
    var_produto.to_excel(writer, index=False, sheet_name='VaR_segmento', startrow=30, startcol=1)

    worksheet = writer.sheets['VaR_segmento']
    worksheet.set_column('E:E', 12, numero_float)
    worksheet.set_column('F:F', 12, numero_float)
    worksheet.set_column('C:C', 12, percent)

    writer.save()


    # Formatação dos relatórios
    abas_n_existentes = pd.DataFrame(columns=['cnpj', 'quadro'])
    abas_n_existentes_k = pd.DataFrame(columns=['cnpj', 'quadro'])

    wb = openpyxl.load_workbook(end + 'relatorio_quadro_operacoes.xlsx')

    #####FORMATOS
    # Fonte
    fontObj1 = Font(name='Calibri', bold=True, size=24, color='404040')
    fontObj2 = Font(name='Calibri', bold=False, size=11, color='404040')

    # Borda
    borderObj1 = Border(bottom=Side(border_style='double'), top=Side(border_style='thin'))
    borderObj2 = Border()

    # Cor
    colorObj1 = PatternFill(patternType='solid', fgColor=Color('FFE600'))

    # Alinhamento
    alinObj1 = Alignment(vertical='center', horizontal='center')
    alinObj2 = Alignment(vertical='center', horizontal='left')

    #####Quadro de Operações

    if contador_qo == 1:

        sheet1 = wb.get_sheet_by_name('Dados Institucionais')
        #
        #
        # Retira as gridlines
        sheet1.sheet_view.showGridLines = False

        # Formatação tamanho das linhas
        sheet1.row_dimensions[1].height = 90

        # Formata cor da fonte de todas as células
        for row in sheet1.range('B2:C20'):
            for cell in row:
                cell.font = fontObj2

        # Formata o título
        sheet1.merge_cells('B2:C2')
        sheet1['B2'] = 'Dados Institucionais'
        sheet1['B2'].font = fontObj1
        sheet1['B2'].alignment = alinObj2

        for row in sheet1.range('B2:C2'):
            for cell in row:
                cell.border = borderObj1

                # Cria a parte de informações institucionais e resumo do relatório de crédito
        sheet1['B4'] = 'Nome'
        sheet1['C4'] = header_nome
        sheet1['B5'] = 'CNPJ'
        sheet1['C5'] = cnpj_fundo
        sheet1['B6'] = 'Administrador'
        sheet1['C6'] = administrador
        sheet1['B7'] = 'Gestor'
        sheet1['C7'] = gestor
        # sheet1['B4']='Nível de Confiança'
        # sheet1['C4']=percentil
        # sheet1['B5']='PL do fundo (R$)'
        # sheet1['C5']=pl_info
        # sheet1['B6']='PL de crédito (R$)'
        # sheet1['C6']=pl_credito
        # sheet1['B7']='Duration - Crédito (DU)'
        # sheet1['C7']=duration_carteira
        # sheet1['B8']='% de crédito privado'
        # sheet1['C8']=pl_credito/pl_info
        # sheet1['B9']='Perda Esperada (em relação ao PL crédito)'
        # sheet1['C9']=pe
        # sheet1['B10']='Percentil da Perda ('+str(percentil)+')'
        # sheet1['C10']=percentil_conf
        # sheet1['B11']='Rating médio da carteira'
        # sheet1['C11']=rtg_lp
        # sheet1['B12']='Perda Esperada (em relação ao PL)'
        # sheet1['C12']=pe_carteira


        # Formatação tamanho das colunas
        sheet1.column_dimensions['A'].width = 2
        sheet1.column_dimensions['B'].width = 15
        sheet1.column_dimensions['C'].width = 100

        ###Quadro de operações

        sheet1 = wb.get_sheet_by_name('Quadro de Operações')

        # Retira as gridlines
        sheet1.sheet_view.showGridLines = False

        # Formata o alinhamento de todas as células da tabela
        for row in sheet1.range('B4:U10000'):
            for cell in row:
                cell.alignment = alinObj1
                cell.font = fontObj2

        # Formatação tamanho das linhas
        sheet1.row_dimensions[1].height = 90

        # Formata o título
        sheet1.merge_cells('B2:V2')
        sheet1['B2'] = 'Quadro de Operações'
        sheet1['B2'].font = fontObj1

        for row in sheet1.range('B2:V2'):
            for cell in row:
                cell.border = borderObj1


                # Formata os headers da tabela
        for row in sheet1.range('B4:V4'):
            for cell in row:
                cell.fill = colorObj1
                cell.border = borderObj2

        # Formatação tamanho das colunas
        sheet1.column_dimensions['A'].width = 2
        sheet1.column_dimensions['B'].width = 20
        sheet1.column_dimensions['C'].width = 20
        sheet1.column_dimensions['D'].width = 30
        sheet1.column_dimensions['E'].width = 20
        sheet1.column_dimensions['F'].width = 20
        sheet1.column_dimensions['G'].width = 20
        sheet1.column_dimensions['H'].width = 100
        sheet1.column_dimensions['I'].width = 25
        sheet1.column_dimensions['J'].width = 15
        sheet1.column_dimensions['K'].width = 25
        sheet1.column_dimensions['L'].width = 15
        sheet1.column_dimensions['M'].width = 25
        sheet1.column_dimensions['N'].width = 25
        sheet1.column_dimensions['O'].width = 15
        sheet1.column_dimensions['P'].width = 15
        sheet1.column_dimensions['Q'].width = 15
        sheet1.column_dimensions['R'].width = 30
        sheet1.column_dimensions['S'].width = 150
        sheet1.column_dimensions['T'].width = 30
        sheet1.column_dimensions['U'].width = 160
        sheet1.column_dimensions['V'].width = 15
    else:
        abas_n_existentes_k['cnpj'] = cnpj
        abas_n_existentes_k['quadro'] = 'operacoes'
        abas_n_existentes = abas_n_existentes.append(abas_n_existentes_k)

    #####Despesas
    if contador_desp == 1:

        sheet2 = wb.get_sheet_by_name('Despesas')

        # Retira as gridlines
        sheet2.sheet_view.showGridLines = False

        # Formata o alinhamento de todas as células da tabela
        for row in sheet2.range('B4:C30'):
            for cell in row:
                cell.alignment = alinObj1
                cell.font = fontObj2


                # Formatação tamanho das linhas
        sheet2.row_dimensions[1].height = 90

        # Formata o título
        sheet2.merge_cells('B2:C2')
        sheet2['B2'] = 'Despesas'
        sheet2['B2'].font = fontObj1

        for row in sheet2.range('B2:C2'):
            for cell in row:
                cell.border = borderObj1


                # Formata os headers da tabela
        for row in sheet2.range('B4:C4'):
            for cell in row:
                cell.fill = colorObj1
                cell.border = borderObj2

        # Formatação tamanho das colunas
        sheet2.column_dimensions['A'].width = 2
        sheet2.column_dimensions['B'].width = 50
        sheet2.column_dimensions['C'].width = 15
    else:
        abas_n_existentes_k['cnpj'] = cnpj
        abas_n_existentes_k['quadro'] = 'despesas'
        abas_n_existentes = abas_n_existentes.append(abas_n_existentes_k)

    #####Rating por Contraparte
    if contador_contraparte == 1:

        sheet3 = wb.get_sheet_by_name('Rating por Contraparte')

        # Retira as gridlines
        sheet3.sheet_view.showGridLines = False

        # Formata o alinhamento de todas as células da tabela
        for row in sheet3.range('B4:G2000'):
            for cell in row:
                cell.alignment = alinObj1
                cell.font = fontObj2

                # Formatação tamanho das linhas
        sheet3.row_dimensions[1].height = 90

        # Formata o título
        sheet3.merge_cells('B2:G2')
        sheet3['B2'] = 'Rating por Contraparte'
        sheet3['B2'].font = fontObj1

        for row in sheet3.range('B2:G2'):
            for cell in row:
                cell.border = borderObj1


                # Formata os headers da tabela
        for row in sheet3.range('B4:G4'):
            for cell in row:
                cell.fill = colorObj1
                cell.border = borderObj2

        # Formatação tamanho das colunas
        sheet3.column_dimensions['A'].width = 2
        sheet3.column_dimensions['B'].width = 20
        sheet3.column_dimensions['C'].width = 100
        sheet3.column_dimensions['D'].width = 15
        sheet3.column_dimensions['E'].width = 20
        sheet3.column_dimensions['F'].width = 20
        sheet3.column_dimensions['G'].width = 20
    else:
        abas_n_existentes_k['cnpj'] = cnpj
        abas_n_existentes_k['quadro'] = 'contraparte'
        abas_n_existentes = abas_n_existentes.append(abas_n_existentes_k)

    #####Rating por Isin
    if contador_isin == 1:

        sheet4 = wb.get_sheet_by_name('Rating por Isin')

        # Retira as gridlines
        sheet4.sheet_view.showGridLines = False

        # Formata o alinhamento de todas as células da tabela
        for row in sheet4.range('B4:M5000'):
            for cell in row:
                cell.alignment = alinObj1
                cell.font = fontObj2


                # Formatação tamanho das linhas
        sheet4.row_dimensions[1].height = 90

        # Formata o título
        sheet4.merge_cells('B2:M2')
        sheet4['B2'] = 'Rating por Contraparte'
        sheet4['B2'].font = fontObj1

        for row in sheet4.range('B2:M2'):
            for cell in row:
                cell.border = borderObj1


                # Formata os headers da tabela
        for row in sheet4.range('B4:M4'):
            for cell in row:
                cell.fill = colorObj1
                cell.border = borderObj2

        # Formatação tamanho das colunas
        sheet4.column_dimensions['A'].width = 2
        sheet4.column_dimensions['B'].width = 20
        sheet4.column_dimensions['C'].width = 100
        sheet4.column_dimensions['D'].width = 15
        sheet4.column_dimensions['E'].width = 15
        sheet4.column_dimensions['F'].width = 30
        sheet4.column_dimensions['G'].width = 20
        sheet4.column_dimensions['H'].width = 30
        sheet4.column_dimensions['I'].width = 30
        sheet4.column_dimensions['J'].width = 30
        sheet4.column_dimensions['K'].width = 30
        sheet4.column_dimensions['L'].width = 30
        sheet4.column_dimensions['M'].width = 30
    else:
        abas_n_existentes_k['cnpj'] = cnpj
        abas_n_existentes_k['quadro'] = 'isin'
        abas_n_existentes = abas_n_existentes.append(abas_n_existentes_k)

    #####Quadro Geral
    sheet5 = wb.get_sheet_by_name('Quadro Geral')

    # Retira as gridlines
    sheet5.sheet_view.showGridLines = False

    # Formata o alinhamento de todas as células da tabela
    for row in sheet5.range('B4:G100'):
        for cell in row:
            cell.alignment = alinObj1
            cell.font = fontObj2


    # Formatação tamanho das linhas
    sheet5.row_dimensions[1].height = 90

    # ---Quadro Geral - Quadro Resumo
    #  Formata o título da tabela
    sheet5 = wb.get_sheet_by_name('Quadro Geral')
    sheet5.merge_cells('B2:G2')
    sheet5['B2'] = 'Quadro Resumo'
    sheet5['B2'].font = fontObj1
    sheet5['B2'].alignment = alinObj2

    for row in sheet5.range('B2:G2'):
        for cell in row:
            cell.border = borderObj1

    # Formata os headers da tabela
    for row in sheet5.range('B4:G4'):
        for cell in row:
            cell.fill = colorObj1
            cell.border = borderObj2

    # Formata DV100% para porcentagem
    for row in sheet5.range('F5:F11'):
        for cell in row:
            cell.number_format = '0.00%'

    # Formata participação% para porcentagem
    for row in sheet5.range('G5:G11'):
        for cell in row:
            cell.number_format = '0.00%'

    # ---Quadro Geral - Títulos Indexados
    if contador_titulos == 1:

        # Formata o título da tabela
        b = 'B' + str(3 + n_linhas_expo + 3)
        sheet5[b] = 'Títulos Indexados'
        sheet5[b].font = fontObj1
        sheet5[b].alignment = alinObj2

        g = 'G' + str(3 + n_linhas_expo + 3)
        for row in sheet5.range(b + ':' + g):
            for cell in row:
                cell.border = borderObj1


                # Formata os headers da tabela
        b = 'B' + str(5 + n_linhas_expo + 3)
        g = 'G' + str(5 + n_linhas_expo + 3)
        for row in sheet5.range(b + ':' + g):
            for cell in row:
                cell.fill = colorObj1
                cell.border = borderObj2

        # Formata o de participação para porcentagem
        d1 = 'D' + str(6 + n_linhas_expo + 3)
        d2 = 'D' + str(n_cat_prod)
        for row in sheet5.range(d1 + ':' + d2):
            for cell in row:
                cell.number_format = '0.00%'

    else:
        abas_n_existentes_k['cnpj'] = cnpj
        abas_n_existentes_k['quadro'] = 'titulos'
        abas_n_existentes = abas_n_existentes.append(abas_n_existentes_k)

    # ---Quadro Geral - Grupos de Produtos
    if contador_grupo_prod == 1:

        # Formata o título da tabela
        b = 'B' + str(n_cat_prod + 2)
        sheet5[b] = 'Grupos de Produtos'
        sheet5[b].font = fontObj1
        sheet5[b].alignment = alinObj2

        g = 'G' + str(n_cat_prod + 2)
        for row in sheet5.range(b + ':' + g):
            for cell in row:
                cell.border = borderObj1


                # Formata os headers da tabela
        b = 'B' + str(n_cat_prod + 2 + 2)
        g = 'G' + str(n_cat_prod + 2 + 2)
        for row in sheet5.range(b + ':' + g):
            for cell in row:
                cell.fill = colorObj1
                cell.border = borderObj2

        # Formata MTM% para porcentagem
        d1 = 'D' + str(n_cat_prod + 2 + 3)
        d2 = 'D' + str(n_cat_prod + 2 + 18)
        for row in sheet5.range(d1 + ':' + d2):
            for cell in row:
                cell.number_format = '0.0000%'
    else:
        abas_n_existentes_k['cnpj'] = cnpj
        abas_n_existentes_k['quadro'] = 'grupo_produto'
        abas_n_existentes = abas_n_existentes.append(abas_n_existentes_k)

    # ---Quadro Geral - Ações
    if contador_acoes == 1:
        # Formata o título da tabela
        b = 'B' + str(n_acoes + 5)
        sheet5[b] = 'Ações'
        sheet5[b].font = fontObj1
        sheet5[b].alignment = alinObj2

        g = 'G' + str(n_acoes + 5)
        for row in sheet5.range(b + ':' + g):
            for cell in row:
                cell.border = borderObj1


                # Formata os headers da tabela
        b = 'B' + str(n_acoes + 5 + 2)
        g = 'G' + str(n_acoes + 5 + 2)
        for row in sheet5.range(b + ':' + g):
            for cell in row:
                cell.fill = colorObj1
                cell.border = borderObj2

        # Formata Participação (%) para porcentagem
        d1 = 'D' + str(n_acoes + 5 + 3)
        d2 = 'D' + str(n_acoes + 5 + 10)
        for row in sheet5.range(d1 + ':' + d2):
            for cell in row:
                cell.number_format = '0.00%'
    else:
        abas_n_existentes_k['cnpj'] = cnpj
        abas_n_existentes_k['quadro'] = 'acoes'
        abas_n_existentes = abas_n_existentes.append(abas_n_existentes_k)

    # Formatação tamanho das colunas
    sheet5.column_dimensions['A'].width = 2
    sheet5.column_dimensions['B'].width = 35
    sheet5.column_dimensions['C'].width = 20
    sheet5.column_dimensions['D'].width = 20
    sheet5.column_dimensions['E'].width = 20
    sheet5.column_dimensions['F'].width = 20
    sheet5.column_dimensions['G'].width = 20

    #####Descasamento PRE
    if 'PRE' in contador_descasamento:

        sheet6 = wb.get_sheet_by_name('descasamento_PRE')

        # Retira as gridlines
        sheet6.sheet_view.showGridLines = False

        # Formata o alinhamento de todas as células da tabela
        for row in sheet6.range('B4:D30'):
            for cell in row:
                cell.alignment = alinObj1
                cell.font = fontObj2


        # Formatação tamanho das linhas
        sheet6.row_dimensions[1].height = 90

        # Formata o título
        sheet6.merge_cells('B2:D2')
        sheet6['B2'] = 'Descasamento - PRE'
        sheet6['B2'].font = fontObj1

        for row in sheet6.range('B2:D2'):
            for cell in row:
                cell.border = borderObj1


        # Formata os headers da tabela
        for row in sheet6.range('B4:D4'):
            for cell in row:
                cell.fill = colorObj1
                cell.border = borderObj2

        # Formata vértice da tabela
        for row in sheet6.range('C5:C30'):
            for cell in row:
                cell.number_format = '0'

        # Formatação tamanho das colunas
        sheet6.column_dimensions['A'].width = 2
        sheet6.column_dimensions['B'].width = 20
        sheet6.column_dimensions['C'].width = 20
        sheet6.column_dimensions['D'].width = 20
    else:
        abas_n_existentes_k['cnpj'] = cnpj
        abas_n_existentes_k['quadro'] = 'descasamento_PRE'
        abas_n_existentes = abas_n_existentes.append(abas_n_existentes_k)

    #####Descasamento IPCA
    if 'IPCA' in contador_descasamento:

        sheet7 = wb.get_sheet_by_name('descasamento_IPCA')

        # Retira as gridlines
        sheet7.sheet_view.showGridLines = False

        # Formata o alinhamento de todas as células da tabela
        for row in sheet7.range('B4:D30'):
            for cell in row:
                cell.alignment = alinObj1
                cell.font = fontObj2


                # Formatação tamanho das linhas
        sheet7.row_dimensions[1].height = 90

        # Formata o título
        sheet7.merge_cells('B2:D2')
        sheet7['B2'] = 'Descasamento - IPCA'
        sheet7['B2'].font = fontObj1

        for row in sheet7.range('B2:D2'):
            for cell in row:
                cell.border = borderObj1


                # Formata os headers da tabela
        for row in sheet7.range('B4:D4'):
            for cell in row:
                cell.fill = colorObj1
                cell.border = borderObj2

        # Formata vértice da tabela
        for row in sheet7.range('C5:C30'):
            for cell in row:
                cell.number_format = '0'

        # Formatação tamanho das colunas
        sheet7.column_dimensions['A'].width = 2
        sheet7.column_dimensions['B'].width = 20
        sheet7.column_dimensions['C'].width = 20
        sheet7.column_dimensions['D'].width = 20
    else:
        abas_n_existentes_k['cnpj'] = cnpj
        abas_n_existentes_k['quadro'] = 'descasamento_IPCA'
        abas_n_existentes = abas_n_existentes.append(abas_n_existentes_k)

    #####Descasamento IGPM
    if 'IGPM' in contador_descasamento:

        sheet8 = wb.get_sheet_by_name('descasamento_IGPM')

        # Retira as gridlines
        sheet8.sheet_view.showGridLines = False

        # Formata o alinhamento de todas as células da tabela
        for row in sheet8.range('B4:D30'):
            for cell in row:
                cell.alignment = alinObj1
                cell.font = fontObj2


                # Formatação tamanho das linhas
        sheet8.row_dimensions[1].height = 90

        # Formata o título
        sheet8.merge_cells('B2:D2')
        sheet8['B2'] = 'Descasamento - IGPM'
        sheet8['B2'].font = fontObj1

        for row in sheet8.range('B2:D2'):
            for cell in row:
                cell.border = borderObj1


                # Formata os headers da tabela
        for row in sheet8.range('B4:D4'):
            for cell in row:
                cell.fill = colorObj1
                cell.border = borderObj2

        # Formata vértice da tabela
        for row in sheet8.range('C5:C30'):
            for cell in row:
                cell.number_format = '0'

        # Formatação tamanho das colunas
        sheet8.column_dimensions['A'].width = 2
        sheet8.column_dimensions['B'].width = 20
        sheet8.column_dimensions['C'].width = 20
        sheet8.column_dimensions['D'].width = 20
    else:
        abas_n_existentes_k['cnpj'] = cnpj
        abas_n_existentes_k['quadro'] = 'descasamento_IGPM'
        abas_n_existentes = abas_n_existentes.append(abas_n_existentes_k)

    #####Resumo - Visão Consolidada
    sheet9 = wb.get_sheet_by_name('Resumo - Visão Consolidada')

    # Retira as gridlines
    sheet9.sheet_view.showGridLines = False

    # Formata o alinhamento de todas as células da tabela
    for row in sheet9.range('B4:G200'):
        for cell in row:
            cell.alignment = alinObj1
            cell.font = fontObj2

    # Formatação tamanho das linhas
    sheet9.row_dimensions[1].height = 90

    # ---Resumo - Visão Consolidada - Parâmetros do Modelo
    # Formata o título da tabela
    sheet9.merge_cells('B2:G2')
    sheet9['B2'] = 'Parâmetros do Modelo'
    sheet9['B2'].font = fontObj1
    sheet9['B2'].alignment = alinObj2

    for row in sheet9.range('B2:G2'):
        for cell in row:
            cell.border = borderObj1


    # Formata os headers da tabela Parametros do Modelo
    for row in sheet9.range('B4:G4'):
        for cell in row:
            cell.fill = colorObj1
            cell.border = borderObj2

    # Ajusta formato de número para perc - Value at Risk (%)
    sheet9['D5'].number_format = '0'

    # ---Resumo - Visão Consolidada - VaR Total
    # Formata o título da tabela
    sheet9.merge_cells('B7:G7')
    sheet9['B7'] = 'Var - Resumo: Total'
    sheet9['B7'].font = fontObj1
    sheet9['B7'].alignment = alinObj2

    for row in sheet9.range('B7:G7'):
        for cell in row:
            cell.border = borderObj1

    # Formata os headers da tabela Parametros do Modelo
    for row in sheet9.range('B9:G9'):
        for cell in row:
            cell.fill = colorObj1
            cell.border = borderObj2

    # Ajusta formato de número para perc - Value at Risk (%)
    sheet9['D10'].number_format = '0.0000%'
    sheet9['F10'].number_format = '0.0000%'

    # ---Resumo - Visão Consolidada - VaR Segmento
    # Formata o título da tabela
    sheet9.merge_cells('B12:G12')
    sheet9['B12'] = 'Var - Resumo: Segmento'
    sheet9['B12'].font = fontObj1
    sheet9['B12'].alignment = alinObj2

    for row in sheet9.range('B12:G12'):
        for cell in row:
            cell.border = borderObj1

    # Formata os headers da tabela VaR Segmento
    for row in sheet9.range('B14:G14'):
        for cell in row:
            cell.fill = colorObj1
            cell.border = borderObj2

    # Formata %VaR para porcentagem
    c1 = 'C' + str(15)
    d2 = 'D' + str(15 + 5)
    for row in sheet9.range(c1 + ':' + d2):
        for cell in row:
            cell.number_format = '0.0000%'

    # ---Resumo - Visão Consolidada - VaR Componente e Marginal
    # Formata o título da tabela
    sheet9.merge_cells('B21:G21')
    sheet9['B21'] = 'Var Componente e Marginal por Fator de Risco'
    sheet9['B21'].font = fontObj1
    sheet9['B21'].alignment = alinObj2

    for row in sheet9.range('B21:G21'):
        for cell in row:
            cell.border = borderObj1


    # Formata os headers da tabela Parametros do Modelo
    for row in sheet9.range('B23:G23'):
        for cell in row:
            cell.fill = colorObj1
            cell.border = borderObj2

    # Formata VaR Marginal para decimal
    f1 = 'F' + str(24)
    g2 = 'G' + str(24 + 100)
    for row in sheet9.range(f1 + ':' + g2):
        for cell in row:
            cell.number_format = '0.0000'

    # Formatação tamanho das colunas
    sheet9.column_dimensions['A'].width = 2
    sheet9.column_dimensions['B'].width = 25
    sheet9.column_dimensions['C'].width = 25
    sheet9.column_dimensions['D'].width = 30
    sheet9.column_dimensions['E'].width = 30
    sheet9.column_dimensions['F'].width = 30
    sheet9.column_dimensions['G'].width = 30

    #####Resumo - Var Segmento

    sheet10 = wb.get_sheet_by_name('VaR_segmento')

    # Retira as gridlines
    sheet10.sheet_view.showGridLines = False

    # Formata o alinhamento de todas as células da tabela
    for row in sheet10.range('B4:H3000'):
        for cell in row:
            cell.alignment = alinObj1
            cell.font = fontObj2


    # Formatação tamanho das linhas
    sheet10.row_dimensions[1].height = 90

    # ---VaR Segmento - Var por segmento
    # Formata o título da tabela
    sheet10.merge_cells('B2:H2')
    sheet10['B2'] = 'VaR por Segmento'
    sheet10['B2'].font = fontObj1
    sheet10['B2'].alignment = alinObj2

    for row in sheet10.range('B2:H2'):
        for cell in row:
            cell.border = borderObj1

    # Formata os headers da tabela Parametros do Modelo
    for row in sheet10.range('B4:H4'):
        for cell in row:
            cell.fill = colorObj1
            cell.border = borderObj2

    # Formata %VaR para porcentagem
    for row in sheet10.range('C5:F19'):
        for cell in row:
            cell.number_format = '0.0000%'

    # ---VaR Segmento - VaR por Produto
    # Formata o título da tabela
    sheet10.merge_cells('B29:H29')
    sheet10['B29'] = 'VaR por Produto'
    sheet10['B29'].font = fontObj1
    sheet10['B29'].alignment = alinObj2

    for row in sheet10.range('B29:H29'):
        for cell in row:
            cell.border = borderObj1

    # Formata os headers da tabela Parametros do Modelo
    for row in sheet10.range('B31:H31'):
        for cell in row:
            cell.fill = colorObj1
            cell.border = borderObj2

    # Formata VaR Marginal para quatro decimais
    for row in sheet10.range('G32:H3000'):
        for cell in row:
            cell.number_format = '0.0000'

    # Formatação tamanho das colunas
    sheet10.column_dimensions['A'].width = 2
    sheet10.column_dimensions['B'].width = 30
    sheet10.column_dimensions['C'].width = 30
    sheet10.column_dimensions['D'].width = 30
    sheet10.column_dimensions['E'].width = 30
    sheet10.column_dimensions['F'].width = 30
    sheet10.column_dimensions['G'].width = 30
    sheet10.column_dimensions['H'].width = 30

    wb.save(end + cnpj + "_" + dt_base + 'relatorio_de_risco.xlsx')

    abas_n_existentes['data_bd'] = horario_bd
    pd.io.sql.to_sql(abas_n_existentes, name='abas', con=connection, if_exists="append", flavor='mysql', index=0,chunksize=5000)

    #Fecha conexão
    connection.close()