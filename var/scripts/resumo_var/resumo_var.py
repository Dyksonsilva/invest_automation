import pandas as pd
import pymysql as db
import openpyxl
import logging

from openpyxl.styles import Font, Color, Border, Side, PatternFill, Alignment
from dependencias.Metodos.funcoes_auxiliares import full_path_from_database
from dependencias.Metodos.funcoes_auxiliares import get_global_var
from dependencias.Metodos.funcoes_auxiliares import get_data_ultimo_dia_util_mes_anterior

global header_id_carteira
global header_id_carteira_fundos
global header_nome
global cnpj_fundo
global id_relatorio_qo

logger = logging.getLogger(__name__)
end = full_path_from_database("get_output_var")+'Relatorios/'
cnpj_hdi=get_global_var("cnpj_hdi")
dt_base = get_data_ultimo_dia_util_mes_anterior()
dt_base = dt_base[0]+'-'+dt_base[1]+'-'+dt_base[2]
dt_base = '2016-11-30'

logger.info("Conectando no Banco de dados")
connection=db.connect('localhost',user='root',passwd='root', db='projeto_inv', use_unicode=True, charset="utf8")
logger.info("Conexão com DB executada com sucesso")

tabela_resumo_var = pd.DataFrame(columns=['cnpj','nome','pl','var_normal','var_estressado','perc_var_norm','perc_var_stress'])
tabela_resumo_credito = pd.DataFrame(columns=['cnpj','nome','pl','pl_cred','perc_esp','rtg_medio'])

#Gera lista dos CNPJs na carteira HDI
#Busca lista dos fundos de primeiro nivel na carteira da HDI

query='select * from projeto_inv.xml_header where cnpjcpf="' + cnpj_hdi +'" and dtposicao='+'"'+dt_base+'";'
df=pd.read_sql(query, con=connection)

if len(df)==0:
    query='select * from projeto_inv.xml_header where cnpj="' + cnpj_hdi +'" and dtposicao='+'"'+dt_base+'";'
    df = pd.read_sql(query, con=connection)

df=df.sort(['cnpj', 'cnpjcpf','data_bd'], ascending=[True, True, False])
df=df.drop_duplicates(subset=['cnpj', 'cnpjcpf'], take_last=False)
df=df.reset_index(level=None, drop=False, inplace=False, col_level=0, col_fill='')
del df['index']

header_id_carteira_fundos=df.get_value(0,'header_id').astype(str) #Utiliza o header da carteira da HDI como chave para a query da lista    


lista_query='SELECT cnpj from projeto_inv.lista_fundos where data_bd=(select max(data_bd) from projeto_inv.lista_fundos where header_id="'+header_id_carteira_fundos+'");'
lista_cnpj=pd.read_sql(lista_query, con=connection)
logger.info("Leitura do banco de dados executada com sucesso")

lista=lista_cnpj['cnpj'].tolist()
lista_query_fidc='SELECT * from projeto_inv.lista_fidc where data_bd=(select max(data_bd) from projeto_inv.lista_fidc where header_id="'+header_id_carteira_fundos+'");'
lista_fidc=pd.read_sql(lista_query_fidc, con=connection)
logger.info("Leitura do banco de dados executada com sucesso")

lista_fidc=lista_fidc['cnpj'].tolist()

for cnpj in lista:

    #Busca do id_relatorio_quaid419
    query='select * from projeto_inv.xml_header where cnpjcpf="' + cnpj +'" and dtposicao='+'"'+dt_base+'";'
    df=pd.read_sql(query, con=connection)
    logger.info("Leitura do banco de dados executada com sucesso")
    if len(df) == 0:
        query = 'select * from projeto_inv.xml_header where cnpj="' + cnpj +'" and dtposicao='+'"'+dt_base+'";'
        df = pd.read_sql(query, con=connection)
        logger.info("Leitura do banco de dados executada com sucesso")
    df = df.sort(['cnpj', 'cnpjcpf','data_bd'], ascending=[True, True, False])
    df = df.drop_duplicates(subset=['cnpj', 'cnpjcpf'], take_last=False)
    df = df.reset_index(level=None, drop=False, inplace=False, col_level=0, col_fill='')
    del df['index']

    header_id_carteira=df.get_value(0,'header_id').astype(str)
    header_nome=df.get_value(0,'nome')
    cnpj_fundo=cnpj

    #quadro de operaçoes
    query='select a.* from projeto_inv.xml_quadro_operacoes a right join (select header_id, max(data_bd) as data_bd from projeto_inv.xml_quadro_operacoes where header_id='+header_id_carteira+' group by 1) b on a.header_id=b.header_id and a.data_bd=b.data_bd;'
    qo=pd.read_sql(query, con=connection)
    logger.info("Leitura do banco de dados executada com sucesso")
    id_relatorio_qo=str(qo['id_relatorio_qo'][0])
    if cnpj=='29980158000157':
        id_relatorio_qo_hdi=id_relatorio_qo
    
#    id_relatorio_qo = "50"
    query='select a.* from projeto_inv.quaid_419 a right join (select id_relatorio_qo, tipo_relatorio, max(data_bd) as data_bd from projeto_inv.quaid_419 where id_relatorio_qo='+id_relatorio_qo+' and tipo_relatorio="G" group by 1,2) b on a.id_relatorio_qo=b.id_relatorio_qo and a.tipo_relatorio=b.tipo_relatorio and a.data_bd=b.data_bd;'
    quaid_419=pd.read_sql(query, con=connection)
    logger.info("Leitura do banco de dados executada com sucesso")
    quaid_419=quaid_419[quaid_419.EMFMULTIPLOFATOR==0]    #Se tem linha duplicada, mais de um fator de risco, pega apenas a referente ao indexador principal
    id_quaid_419_df=quaid_419[['id_relatorio_quaid419']]
    id_relatorio_quaid419=id_quaid_419_df.get_value(0,'id_relatorio_quaid419').astype(str)
    #print(cnpj,id_relatorio_qo,id_relatorio_quaid419)

    #Patrimônio líquido
    #Informações de PL
    pl_qo=qo[qo.produto!='Futuro']
    pl_info=sum(pl_qo['mtm_info'])

    #Risco de Mercado - VaR
    query='select a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, tipo_segmento, norm_stress, max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419='+str(id_relatorio_quaid419)+' and tipo_var="Total" and vertice="Total" and tipo_alocacao="Total" and tipo_segmento="Total" and norm_stress = "normal" group by 1,2,3,4,5,6) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.tipo_alocacao=b.tipo_alocacao and a.tipo_segmento=b.tipo_segmento and a.norm_stress = b.norm_stress and a.data_bd=b.data_bd; '
    resumo_var_normal=pd.read_sql(query, con=connection)
    logger.info("Leitura do banco de dados executada com sucesso")
    resumo_var_normal = resumo_var_normal[['id_relatorio_quaid419','var']].copy()
    resumo_var_normal = resumo_var_normal.rename(columns={'var':'var_normal'})

    query='select a.* from projeto_inv.var a right join (select id_relatorio_quaid419, tipo_var, vertice, tipo_alocacao, tipo_segmento, norm_stress, max(data_bd) as data_bd from projeto_inv.var where id_relatorio_quaid419='+str(id_relatorio_quaid419)+' and tipo_var="Total" and vertice="Total" and tipo_alocacao="Total" and tipo_segmento="Total" and norm_stress = "estressado" group by 1,2,3,4,5,6) b on a.id_relatorio_quaid419=b.id_relatorio_quaid419 and a.tipo_var=b.tipo_var and a.tipo_alocacao=b.tipo_alocacao and a.tipo_segmento=b.tipo_segmento and a.norm_stress = b.norm_stress and a.data_bd=b.data_bd; '
    resumo_var_estressado=pd.read_sql(query, con=connection)
    logger.info("Leitura do banco de dados executada com sucesso")
    resumo_var_estressado = resumo_var_estressado[['id_relatorio_quaid419','var']].copy()
    resumo_var_estressado = resumo_var_estressado.rename(columns={'var':'var_estressado'})

    resumo_var = pd.merge(resumo_var_normal,resumo_var_estressado,left_on=['id_relatorio_quaid419'],right_on=['id_relatorio_quaid419'],how='left')
    resumo_var['cnpj'] = cnpj
    resumo_var['nome'] = header_nome
    resumo_var['pl'] = pl_info
    resumo_var['perc_var_norm'] = resumo_var['var_normal']/pl_info
    resumo_var['perc_var_stress'] = resumo_var['var_estressado']/pl_info
    del resumo_var['id_relatorio_quaid419']

    #Risco de Crédito - Percentual de perda esperada e rating médio da carteira
    query='select a.* from projeto_inv.resumo_credito a right join (select cnpj, max(data_bd) as data_bd from projeto_inv.resumo_credito where cnpj='+cnpj+' group by 1) b on a.cnpj = b.cnpj and a.data_bd=b.data_bd; '
    resumo_credito=pd.read_sql(query, con=connection)
    logger.info("Leitura do banco de dados executada com sucesso")

    #União das tabelas
    tabela_resumo_var = tabela_resumo_var.append(resumo_var)
    if len(resumo_credito)>0:
        resumo_credito['pl'] = pl_info
        tabela_resumo_credito = tabela_resumo_credito.append(resumo_credito)
        
#FIDCS e FIP
#Cria dataframe com a exposicao dos FIDCs (busca a informação no quadro consolidado)
tabela_fidc=pd.DataFrame()

for cnpj in lista_fidc:
    quadro_fidc=pd.read_sql('select mtm_info, cnpjfundo_1nivel, fundo, isin from projeto_inv.xml_quadro_operacoes where data_bd=(select max(data_bd) from projeto_inv.xml_quadro_operacoes where header_id="'+header_id_carteira_fundos+'") and cnpjfundo_1nivel="'+cnpj+'" and header_id="'+header_id_carteira_fundos+'";', con=connection)
    logger.info("Leitura do banco de dados executada com sucesso")
    tabela_fidc=tabela_fidc.append(quadro_fidc)
    
lista_isin=tabela_fidc['isin'].tolist()

#Extrai o valor do id do quadro 419 da HDI
id_quadro419=pd.read_sql('SELECT id_relatorio_quaid419 FROM projeto_inv.quaid_419 where data_bd=(select max(data_bd) from projeto_inv.quaid_419 where id_relatorio_qo="'+id_relatorio_qo_hdi+'" and tipo_relatorio="G") and id_relatorio_qo="'+id_relatorio_qo_hdi+'";', con=connection)
id_quadro419=id_quadro419.get_value(0,'id_relatorio_quaid419').astype(str) 

#Extrai os valores de VaR de cada um dos ISINs presentes na lista_isin
var_fidc=pd.DataFrame()

for isin in lista_isin:
    var_fidc_bd=pd.read_sql('SELECT var, vertice from projeto_inv.var where data_bd=(select max(data_bd) from projeto_inv.var where id_relatorio_quaid419="'+id_quadro419+'" and norm_stress="normal") and vertice="'+isin+'" and id_relatorio_quaid419="'+id_quadro419+'" and tipo_var="Marginal" and tipo_alocacao="Total" and tipo_segmento="Total" and norm_stress="normal";', con=connection)
    logger.info("Leitura do banco de dados executada com sucesso")
    var_fidc=var_fidc.append(var_fidc_bd)

#Fecha conexão
connection.close()

var_fidc=pd.merge(var_fidc, tabela_fidc, left_on=['vertice'], right_on=['isin'], how='inner')
var_fidc['var']=var_fidc['var']*var_fidc['mtm_info']
var_fidc['VaR Estressado (R$)']=var_fidc['var']*1.14566452
var_fidc['VaR (%PL)']=var_fidc['var']/var_fidc['mtm_info']
var_fidc['VaR Estressado (%PL)']=var_fidc['VaR Estressado (R$)']/var_fidc['mtm_info']
var_fidc=var_fidc.rename(columns={'var':'VaR (R$)', 'mtm_info':'Patrimônio Líquido (R$)', 'fundo':'Nome do Fundo', 'cnpjfundo_1nivel':'CNPJ'})
var_fidc=var_fidc[['CNPJ', 'Nome do Fundo', 'Patrimônio Líquido (R$)', 'VaR (R$)', 'VaR Estressado (R$)', 'VaR (%PL)', 'VaR Estressado (%PL)']].copy()

#Geração do relatorio
writer = pd.ExcelWriter(end+'relatorio_risco_resumo_'+dt_base+'.xlsx', engine='xlsxwriter')
workbook = writer.book   
numero_float = workbook.add_format({'num_format': '#,##0', 'bold': False})
percent = workbook.add_format({'num_format': '0.0000%', 'bold': False})    
percent1 = workbook.add_format({'num_format': '0%', 'bold': False})    

#VaR
var = pd.DataFrame(columns=['CNPJ','Nome do Fundo','Patromônio Líquido (R$)','VaR (R$)','VaR Estressado (R$)','VaR (%PL)','VaR Estressado (%PL)'])
var['CNPJ'] = tabela_resumo_var['cnpj']
var['Nome do Fundo'] = tabela_resumo_var['nome']
var['Patromônio Líquido (R$)'] = tabela_resumo_var['pl']
var['VaR (R$)'] = tabela_resumo_var['var_normal']
var['VaR Estressado (R$)'] = tabela_resumo_var['var_estressado']
var['VaR (%PL)'] = tabela_resumo_var['var_normal']/tabela_resumo_var['pl']
var['VaR Estressado (%PL)'] = tabela_resumo_var['var_estressado']/tabela_resumo_var['pl']

tabela_pl = tabela_resumo_var[['cnpj','pl']].copy()

var.to_excel(writer,index=False, sheet_name='Resumo VaR', startrow =3, startcol=1) 
logger.info("Arquivo salvo com sucesso")

tamanho=len(var)

var_fidc.to_excel(writer,index=False, sheet_name='Resumo VaR', startrow =(7+tamanho), startcol=1)   
logger.info("Arquivo salvo com sucesso")

worksheet = writer.sheets['Resumo VaR']
worksheet.set_column('D:D', 12, numero_float)
worksheet.set_column('E:E', 12, numero_float)
worksheet.set_column('F:F', 12, numero_float)
worksheet.set_column('G:G', 12, percent)
worksheet.set_column('H:H', 12, percent)

#Credito
credito = pd.DataFrame(columns=['CNPJ','Nome do Fundo','Patromônio Líquido Crédito (R$)','Perda Esperada (R$)','Perda Esperada (%PL - Crédito)','Rating Médio do Fundo'])
credito['CNPJ'] = tabela_resumo_credito['cnpj']
credito['Nome do Fundo'] = tabela_resumo_credito['nome']
credito['Patromônio Líquido Crédito (R$)'] = tabela_resumo_credito['pl_cred']
credito['Perda Esperada (R$)'] = tabela_resumo_credito['pl']*tabela_resumo_credito['perc_esp']
credito['Perda Esperada (%PL - Crédito)'] = tabela_resumo_credito['pl']*tabela_resumo_credito['perc_esp']/tabela_resumo_credito['pl_cred']
credito['Rating Médio do Fundo'] = tabela_resumo_credito['rtg_medio']
credito.to_excel(writer,index=False, sheet_name='Resumo Crédito', startrow =3, startcol=1)
logger.info("Arquivo salvo com sucesso")

worksheet = writer.sheets['Resumo Crédito']
worksheet.set_column('D5:E30', 12, numero_float)
worksheet.set_column('F5:F30', 12, percent1)

writer.save()

#Formatação resumo
wb = openpyxl.load_workbook(end+'relatorio_risco_resumo_'+dt_base+'.xlsx')

#####FORMATOS
#Fonte
fontObj1=Font(name='Calibri', bold=True, size =24,color='404040')
fontObj2=Font(name='Calibri', bold=False, size =11,color='404040')
    
#Borda
borderObj1=Border(bottom=Side(border_style='double'),top=Side(border_style='thin'))
borderObj2=Border()

#Cor
colorObj1=PatternFill(patternType='solid', fgColor=Color('FFE600'))

#Alinhamento
alinObj1=Alignment(vertical='center',horizontal='center')
alinObj2=Alignment(vertical='center',horizontal='left')
alinObj3=Alignment(vertical='center',horizontal='right')

sheet1=wb.get_sheet_by_name('Resumo VaR')

#Retira as gridlines
sheet1.sheet_view.showGridLines = False

#Formatação tamanho das linhas
sheet1.row_dimensions[1].height = 90

#Formata cor da fonte de todas as células
for row in sheet1.range('B2:H50'):
    for cell in row:
        cell.font=fontObj2

#Formata o título
sheet1.merge_cells('B2:H2')
sheet1['B2']='Resumo VaR'
sheet1['B2'].font=fontObj1
sheet1['B2'].alignment=alinObj2

sheet1.merge_cells('B'+str(6+tamanho)+':H'+str(6+tamanho))#Tabela FIDC
sheet1['B'+str(6+tamanho)]='Resumo VaR FIDCs'
sheet1['B'+str(6+tamanho)].font=fontObj1
sheet1['B'+str(6+tamanho)].alignment=alinObj2

for row in sheet1.range('B2:H2'):
    for cell in row:
        cell.border=borderObj1    

for row in sheet1.range('B'+str(6+tamanho)+':H'+str(6+tamanho)):#Tabela FIDC
    for cell in row:
        cell.border=borderObj1    

#Formata os headers da tabela
for row in sheet1.range('B4:H4'):
    for cell in row:
        cell.fill=colorObj1
        cell.border=borderObj2
        
for row in sheet1.range('B'+str(8+tamanho)+':H'+str(8+tamanho)):#Tabela FIDC
    for cell in row:
        cell.fill=colorObj1
        cell.border=borderObj2

#Formatação tamanho das colunas
sheet1.column_dimensions['A'].width = 2
sheet1.column_dimensions['B'].width = 20
sheet1.column_dimensions['C'].width = 70
sheet1.column_dimensions['D'].width = 30
sheet1.column_dimensions['E'].width = 20
sheet1.column_dimensions['F'].width = 30
sheet1.column_dimensions['G'].width = 25
sheet1.column_dimensions['H'].width = 25

##CRÉDTIO
sheet1=wb.get_sheet_by_name('Resumo Crédito')

#Retira as gridlines
sheet1.sheet_view.showGridLines = False

#Formatação tamanho das linhas
sheet1.row_dimensions[1].height = 90

#Formata cor da fonte de todas as células
for row in sheet1.range('B2:G30'):
   for cell in row:
       cell.font=fontObj2

#Formata o título
sheet1.merge_cells('B2:G2')
sheet1['B2']='Resumo Crédito'
sheet1['B2'].font=fontObj1
sheet1['B2'].alignment=alinObj2

for row in sheet1.range('B2:G2'):
   for cell in row:
       cell.border=borderObj1

#Formata os headers da tabela
for row in sheet1.range('B4:G4'):
   for cell in row:
       cell.fill=colorObj1
       cell.border=borderObj2

#Formatação tamanho das colunas
sheet1.column_dimensions['A'].width = 2
sheet1.column_dimensions['B'].width = 20
sheet1.column_dimensions['C'].width = 70
sheet1.column_dimensions['D'].width = 30
sheet1.column_dimensions['E'].width = 20
sheet1.column_dimensions['F'].width = 30
sheet1.column_dimensions['G'].width = 25

wb.save(end+dt_base+' relatorio_resumo.xlsx')