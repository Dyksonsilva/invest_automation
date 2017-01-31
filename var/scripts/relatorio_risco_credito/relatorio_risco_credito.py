def relatorio_risco_credito():

    import datetime
    import pandas as pd
    import pymysql as db
    import numpy as np
    import openpyxl
    import logging

    from openpyxl.styles import Font, Color, Border, Side, PatternFill, Alignment
    from dependencias.Metodos.funcoes_auxiliares import get_data_ultimo_dia_util_mes_anterior
    from dependencias.Metodos.funcoes_auxiliares import get_global_var
    from dependencias.Metodos.funcoes_auxiliares import full_path_from_database

    global header_id_carteira_fundos

    dt_base = get_data_ultimo_dia_util_mes_anterior()
    dt_base = dt_base[0]+'-'+dt_base[1]+'-'+dt_base[2]
    logger = logging.getLogger(__name__)
    nivel_confianca=0.95
    end = full_path_from_database("get_output_var")+'relatorios/'
    cnpj_hdi=get_global_var("cnpj_hdi")

    logger.info("Conectando no Banco de dados")
    connection=db.connect('localhost',user='root',passwd='root', db='projeto_inv', use_unicode=True, charset="utf8")
    logger.info("Conexão com DB executada com sucesso")

    #Fecha conexão
    connection.close()

    def relatorio_credito(dt_base,cnpj,nivel_confianca,horario_bd):

        global percentil
        global header_id_carteira
        global header_nome
        global cnpj_fundo
        global administrador
        global gestor
        global quadro_oper
        global pl_info
        global tp
        global id_relatorio_qo
        global pl_credito
        global perc_credito
        global pe
        global duration_carteira
        global rtg_lp

        # Create a Pandas Excel writer using XlsxWriter as the engine.
        writer = pd.ExcelWriter(end+cnpj+"_"+dt_base+' relatorio_credito.xlsx', engine='xlsxwriter')
        workbook = writer.book
        numero_float = workbook.add_format({'num_format': '#,##0', 'bold': False})
        percent = workbook.add_format({'num_format': '0.0%', 'bold': False})
        percentil=nivel_confianca

        logger.info("Conectando no Banco de dados")
        connection = db.connect('localhost', user='root', passwd='root', db='projeto_inv', use_unicode=True,
                                charset="utf8")
        logger.info("Conexão com DB executada com sucesso")

        query='select * from projeto_inv.xml_header where cnpjcpf="' + cnpj +'" and dtposicao='+'"'+dt_base+'";'

        df=pd.read_sql(query, con=connection)

        if len(df)==0:
            query='select * from projeto_inv.xml_header where cnpj="' + cnpj +'" and dtposicao='+'"'+dt_base+'";'
            df=pd.read_sql(query, con=connection)

        df=df.sort(['cnpj', 'cnpjcpf','data_bd'], ascending=[True, True, False])
        df=df.drop_duplicates(subset=['cnpj', 'cnpjcpf'], take_last=False)
        df=df.reset_index(level=None, drop=False, inplace=False, col_level=0, col_fill='')
        del df['index']
        header_id_carteira=df.get_value(0,'header_id').astype(str)
        header_nome=df.get_value(0,'nome')
        cnpj_fundo=cnpj
        administrador=df.get_value(0, 'nomeadm')
        gestor=df.get_value(0, 'nomegestor')
        query='select a.* from projeto_inv.xml_quadro_operacoes a right join (select header_id, max(data_bd) as data_bd from projeto_inv.xml_quadro_operacoes where header_id='+header_id_carteira+' group by 1) b on a.header_id=b.header_id and a.data_bd=b.data_bd;'
        quadro_oper=pd.read_sql(query, con=connection)
        pl=quadro_oper[quadro_oper['produto']!='Futuro']
        pl_info=sum(pl['mtm_info'])

        #selecionar crédito
        tp=quadro_oper.loc[quadro_oper['produto'].isin(['título privado','debênture'])]

        #id_relatorio_qo
        id_relatorio_qo=quadro_oper['id_relatorio_qo'].iloc[0]
        id_relatorio_qo_str=str(id_relatorio_qo)

        if len(tp)>0:

            # fidc
            fundos=quadro_oper.loc[quadro_oper['produto'].isin(['fundo'])].copy()
            fundos['fundo_final']=np.where(fundos['fundo_ult_nivel'].isnull(),fundos['fundo'],fundos['fundo_ult_nivel'])

            #INCLUIR FIDCS
            fidc=fundos[fundos['fundo_final'].str.contains('FIDC|DIREITOS|CREDITÓRIO|CREDITORIOS|DIREITOS')]
            tp=tp.append(fidc)
            pl_credito=sum(tp['mtm_info'])
            perc_credito=pl_credito/pl_info

            # incluir tipo de papel
            query= "select distinct codigo_isin, data_bd, tipo_ativo from projeto_inv.bmf_numeraca where tipo_ativo in ('DBS', 'LFI', 'LFN', 'DP', 'C', 'CC','CCC', 'CRI');"
            caracteristica=pd.read_sql(query, con=connection)

            caracteristica=caracteristica.sort(['codigo_isin', 'data_bd'], ascending=[True, False])
            caracteristica=caracteristica.drop_duplicates(subset=['codigo_isin'], take_last=False)
            del caracteristica['data_bd']

            tp=pd.merge(tp, caracteristica, left_on='isin', right_on='codigo_isin', how='left')
            tp.ix[tp.tipo_ativo == 'CC', 'tipo_prod']='CCB'
            tp.ix[tp.tipo_ativo == 'CCC', 'tipo_prod']='CCCB'
            tp.ix[tp.tipo_ativo =='DBS', 'tipo_prod']='Debênture'
            tp.ix[tp.tipo_ativo =='C', 'tipo_prod']='CDB'
            tp.ix[tp.tipo_ativo =='LFI', 'tipo_prod']='LF'
            tp.ix[tp.tipo_ativo =='LFN', 'tipo_prod']='LF'
            tp.ix[tp.tipo_ativo =='DP', 'tipo_prod']='DPGE'
            tp.ix[tp.produto =='fundo', 'tipo_prod']='FIDC'
            tp['tipo_prod']=tp['tipo_prod'].fillna('Outros')
            tp['contraparte']=np.where(tp.produto=='fundo',tp.fundo_final, tp.contraparte)
            tp['cnpj_fundo_final']=np.where((tp.produto=='fundo') & (tp.cnpjfundo_outros.isnull()),tp.cnpjfundo_1nivel,tp.cnpjfundo_outros)
            tp['cnpj']=np.where(tp.produto=='fundo',tp.cnpj_fundo_final,tp.cnpj)
            del tp['codigo_isin']

            #informação de emissor (contraparte)
            query='select distinct nome_emissor, cnpj_emissor, data_criacao_emissor from projeto_inv.bmf_emissor where cnpj_emissor>0;'
            emissor=pd.read_sql(query, con=connection)

            emissor=emissor.sort(['cnpj_emissor', 'data_criacao_emissor'], ascending=[True, False])
            emissor1=emissor.drop_duplicates(subset=['cnpj_emissor'], take_last=False)
            emissor1['cnpj']=emissor1['cnpj_emissor'].astype(float)
            emissor2=emissor1[['cnpj', 'nome_emissor']]

            tp=pd.merge(tp, emissor2, left_on='cnpj', right_on='cnpj', how='left')
            tp['contraparte']=tp['contraparte'].fillna(tp['nome_emissor'])
            del tp['nome_emissor']
            tp['contraparte']=tp['contraparte'].fillna('a')

            # selecionar simulação de perda de crédito
            query='select a.* from projeto_inv.simulacao_credito a right join (select id_relatorio_qo, max(data_bd) as data_bd from projeto_inv.simulacao_credito where id_relatorio_qo = "51" group by 1) b on a.id_relatorio_qo=b.id_relatorio_qo and a.data_bd=b.data_bd;'
            sim=pd.read_sql(query,con=connection)
            sim['perda_perc']=sim['perda_perc'].fillna(0)

            # perda esperada por isin
            perda_isin=sim[['isin', 'perda_perc']].copy()
            perda_isin=perda_isin.groupby(['isin'], as_index=False).mean()
            perda_isin=perda_isin.rename(columns={'perda_perc':'perc_pe'})

            tp_isin_mtm=tp[['cnpj','contraparte','isin','tipo_prod', 'mtm_info']].copy()
            tp_isin_mtm=tp_isin_mtm.groupby(['cnpj','contraparte','isin','tipo_prod'],as_index=False).sum()
            tp_isin_mtm=pd.merge(tp_isin_mtm, perda_isin, left_on='isin', right_on='isin', how='left')
            tp_isin_mtm['perda_esperada'] = tp_isin_mtm['perc_pe']*tp_isin_mtm['mtm_info']
            tp_isin_mtm['perda_esperada']=tp_isin_mtm['perda_esperada'].fillna(0)

            # perda esperada da carteira
            perda=sum(tp_isin_mtm['perda_esperada'])

            pe=perda/pl_credito
            pe_carteira=pe*pl_credito/pl_info

            #composiçao da carteira - por tipo de produto
            prod=tp_isin_mtm[['tipo_prod', 'mtm_info', 'perda_esperada']]
            prod=prod.groupby(['tipo_prod'], as_index=False).sum()
            prod['perc_credito']=prod['mtm_info']/pl_credito
            prod['perc_pl']=prod['mtm_info']/pl_info
            prod['perc_pe']=prod['perda_esperada']/prod['mtm_info']

            del prod['perda_esperada']
            df_aux = pd.DataFrame(columns=['Dados Institucionais'])
            df_aux.to_excel(writer,index=False, sheet_name='Dados Institucionais', startrow =1, startcol=1, header =['Dados Institucionais'])
            prod.to_excel(writer,index=False, sheet_name='Resumo', startrow =14, startcol=1, header =['Tipo de produto', 'Exposição (R$)', '% PL de crédito', '% PL total', '% Perda Esperada'])

            # perda inesperada
            sim1=sim[['isin','numero_simulacao', 'pv_perda', 'pv']]

            isin_mtm=tp[['isin','mtm_info']].copy()
            isin_mtm=isin_mtm.groupby(['isin'],as_index=False).sum()

            sim2=pd.merge(sim1,isin_mtm, left_on='isin', right_on='isin', how='left' )
            sim2['perc_perda']=sim2['pv_perda']/sim2['pv']
            sim2['perda']=sim2['perc_perda']*sim2['mtm_info']
            del sim2['perc_perda']
            del sim2['pv']
            del sim2['pv_perda']
            sim_agr=sim2.groupby(['numero_simulacao'], as_index=False).sum()
            sim_agr['perda']=sim_agr['perda'].fillna(0)
            sim_agr['perc_perda']=sim_agr['perda']/sim_agr['mtm_info']
            sim_agr=sim_agr.sort(columns=['perc_perda'], ascending=True )
            sim_agr=sim_agr.reset_index(level=None, drop=False, inplace=False, col_level=0, col_fill='')

            n_conf= np.floor(len(sim_agr)*percentil)
            percentil_conf=sim_agr['perc_perda'][n_conf-1]

            # Relatório só a MERCADO - cálculo de duration
            query='select a.* from projeto_inv.quaid_419 a right join (select id_relatorio_qo, tipo_relatorio, max(data_bd) as data_bd from projeto_inv.quaid_419 where id_relatorio_qo='+id_relatorio_qo_str+' and tipo_relatorio="R" group by 1,2) b on a.id_relatorio_qo=b.id_relatorio_qo and a.tipo_relatorio=b.tipo_relatorio and a.data_bd=b.data_bd;'

            quaid_419_r=pd.read_sql(query, con=connection)
            quaid_419_r['expo']=np.where(quaid_419_r['TPFOPERADOR']=='-',-1*quaid_419_r['EMFVLREXPRISCO'],quaid_419_r['EMFVLREXPRISCO'] )
            quaid_419_r['prazo_ponderado']=quaid_419_r['expo']*quaid_419_r['EMFPRAZOFLUXO']

            isin_credito=tp[['isin']].copy()
            isin_credito=isin_credito.drop_duplicates(subset=['isin'], take_last=True)

            quaid_419_credito=pd.merge(isin_credito,quaid_419_r, left_on=['isin'], right_on=['EMFCODISIN'], how='left' )

            duration=quaid_419_credito[['EMFCODISIN','expo','prazo_ponderado']].copy()
            duration_ativa=duration[duration.expo>=0].copy()
            duration_ativa_papel=duration_ativa.groupby('EMFCODISIN', as_index=False).sum()
            duration_ativa_papel['duration']=duration_ativa_papel['prazo_ponderado']/duration_ativa_papel['expo']

            #duration carteira
            duration_c=quaid_419_credito[['EMFPRAZOFLUXO','expo']].copy()
            duration_c1=duration_c.groupby(['EMFPRAZOFLUXO'], as_index=False).sum()
            duration_c1['prazo_ponderado']=duration_c1['expo']*duration_c1['EMFPRAZOFLUXO']

            duration_carteira=sum(duration_c1['prazo_ponderado'])/sum(duration_c1['expo'])

            #rating da carteira
            query='select a.rtg, a.prazo, a.pd_acum from projeto_inv.pd_acum a right join (select max(data_bd) as data_bd from projeto_inv.pd_acum) b on a.data_bd=b.data_bd;'
            regua_unica=pd.read_sql(query, con=connection)
            regua_unica['pd_acum_shift']=regua_unica['pd_acum'].shift()
            regua_unica['prazo_shift']=regua_unica['prazo'].shift()
            regua_unica['pd_acum_min']=np.where(regua_unica['prazo_shift']==regua_unica['prazo'],regua_unica['pd_acum_shift'],0)

            del regua_unica['prazo_shift']
            del regua_unica['pd_acum_shift']

            duration_mes=np.floor(duration_carteira/22)

            if duration_mes>60:
                duration_carteira_temp=60
            else:
                duration_carteira_temp=np.floor(duration_mes)

            rtg_carteira=regua_unica[regua_unica['prazo']==duration_carteira_temp].copy()
            rtg_carteira1=rtg_carteira[rtg_carteira['pd_acum_min']<=pe_carteira]
            rtg_carteira2=rtg_carteira1[rtg_carteira1['pd_acum']>pe_carteira]
            rtg_carteira2=rtg_carteira2.reset_index(level=None, drop=False, inplace=False, col_level=0, col_fill='')

            rtg_lp=rtg_carteira2.get_value(0,'rtg')


            # rating por produto
            query='select cod_rtg, rtg from projeto_inv.de_para_rtg a right join (select max(data_bd) as data_bd from projeto_inv.de_para_rtg) b on a.data_bd = b.data_bd;'
            depara=pd.read_sql(query, con=connection)

            #regua mestra
            query='select distinct a.cod_rtg, a.agencia_rtg, a.rtg from projeto_inv.de_para_rtg a right join (select agencia_rtg, max(data_bd) as data_bd from projeto_inv.de_para_rtg where agencia_rtg="regua" group by 1) b on a.agencia_rtg=b.agencia_rtg and a.data_bd = b.data_bd;'
            regua_rtg=pd.read_sql(query, con=connection)
            del regua_rtg['agencia_rtg']

            #rating por isin
            query='select distinct a.isin, a.agencia_tipo_rtg, a.rtg from projeto_inv.rating_isin as a right join (select max(data_bd) as data_bd from projeto_inv.rating_isin where dt_ref= "'+ dt_base +'" ) as b on a.data_bd=b.data_bd;'
            rtg_isin=pd.read_sql(query, con=connection)

            rtg_local=rtg_isin.loc[rtg_isin['agencia_tipo_rtg'].isin(['RTG_MDY_NSR', 'RTG_MDY_NSR_SR_UNSECURED',  'RTG_MDY_NSR_SUBORDINATED','RTG_SP_NATIONAL','RTG_FITCH_NATIONAL_LT', 'RTG_FITCH_NATIONAL', 'RTG_FITCH_NATIONAL_SR_UNSECURED', 'RTG_FITCH_NATL_SUBORDINATED'])]
            rtg_local=pd.merge(rtg_local, depara, left_on='rtg', right_on='rtg', how='left')

            rtg_pior=rtg_local[['isin', 'cod_rtg']].copy()
            rtg_pior=rtg_pior.groupby(['isin'],as_index=False).max()
            rtg_pior=pd.merge(rtg_pior, regua_rtg, left_on='cod_rtg', right_on='cod_rtg', how='left')

            #rating por contraparte
            query='select distinct a.cnpj, a.agencia_tipo_rtg, a.rtg from projeto_inv.rating_contraparte as a right join (select max(data_bd) as data_bd from projeto_inv.rating_contraparte where dt_ref= "'+ dt_base +'" ) as b on a.data_bd=b.data_bd;'
            rtg_c=pd.read_sql(query, con=connection)
            rtg_c_local=rtg_c.loc[rtg_c['agencia_tipo_rtg'].isin(['RTG_MDY_NSR_ISSUER','RTG_SP_NATIONAL_LT_ISSUER_CREDIT','RTG_FITCH_NATIONAL_LT','RTG_FITCH_NATIONAL_SR_UNSECURED' ])]
            rtg_c_local=pd.merge(rtg_c_local, depara, left_on='rtg', right_on='rtg', how='left')
            rtg_c_pior=rtg_c_local[['cnpj', 'cod_rtg']].copy()
            rtg_c_pior=rtg_c_pior.groupby(['cnpj'], as_index=False).max()
            rtg_c_pior=pd.merge(rtg_c_pior, regua_rtg, left_on='cod_rtg', right_on='cod_rtg', how='left')

            #agregar o rtg na base
            tp_cnpj=pd.merge(tp_isin_mtm, rtg_pior, left_on='isin', right_on='isin', how='left')
            tp_cnpj=tp_cnpj.rename(columns={'cod_rtg':'cod_rtg_isin', 'rtg':'rtg_isin'})
            tp_cnpj=pd.merge(tp_cnpj, rtg_c_pior, left_on='cnpj', right_on='cnpj', how='left')
            tp_cnpj=tp_cnpj.rename(columns={'cod_rtg':'cod_rtg_cnpj', 'rtg':'rtg_cnpj'})
            tp_cnpj['cod_rtg']=np.where(tp_cnpj['cod_rtg_isin'].isnull(),tp_cnpj.cod_rtg_cnpj,tp_cnpj.cod_rtg_isin)

            del tp_cnpj['cod_rtg_isin']
            del tp_cnpj['cod_rtg_cnpj']
            del tp_cnpj['rtg_isin']
            del tp_cnpj['rtg_cnpj']

            tp_cnpj=pd.merge(tp_cnpj, regua_rtg, left_on='cod_rtg', right_on='cod_rtg', how='left')

            ##assumir rtg padrão missing: 'Aa3' e cod_rtg=4
            tp_cnpj['cod_rtg']=tp_cnpj['cod_rtg'].fillna(3)
            tp_cnpj['rtg']=tp_cnpj['rtg'].fillna('Aa2')

            sumario_rtg=tp_cnpj[['cod_rtg','rtg', 'mtm_info', 'perda_esperada']]
            sumario_rtg['mtm_info']=sumario_rtg['mtm_info'].fillna(0)
            sumario_rtg['perda_esperada']=sumario_rtg['perda_esperada'].fillna(0)

            sumario=sumario_rtg.groupby(['cod_rtg','rtg'], as_index=False).sum()
            sumario['perc_credito']= sumario['mtm_info']/pl_credito
            sumario['perc_tot']= sumario['mtm_info']/pl_info
            sumario['perc_pe']=sumario['perda_esperada']/sumario['mtm_info']
            del sumario['cod_rtg']
            del sumario['perda_esperada']
            sumario.to_excel(writer,index=False, sheet_name='Resumo', startrow =25, startcol=1, header =['Rating', 'Exposição (R$)', '% PL de crédito', '% PL total', '% Perda Esperada'])

            worksheet = writer.sheets['Resumo']
            worksheet.set_column('D:F', 12, percent)
            worksheet.set_column('C:C', 12, numero_float)

            # 20 maiores exposiçoes por contraparte
            contraparte=tp_cnpj[['cnpj', 'contraparte', 'mtm_info', 'perda_esperada']].copy()
            contraparte=contraparte.groupby(['cnpj', 'contraparte'], as_index=False).sum()
            contraparte['perc_exposicao']=contraparte['mtm_info']/pl_credito
            del contraparte['cnpj']
            contraparte=contraparte.sort(['mtm_info'], ascending=False)
            contraparte=contraparte.reset_index(level=None, drop=False, inplace=False, col_level=0, col_fill='')
            contraparte['perc_pe']=contraparte['perda_esperada']/contraparte['mtm_info']
            del contraparte['perda_esperada']
            del contraparte['index']
            contraparte1=contraparte[0:20]
            contraparte1.to_excel(writer,index=False, sheet_name='Maiores Exposições', startrow =3, startcol=1, header =['Contraparte', 'Exposição', '% PL de crédito', '% Perda Esperada'])

            worksheet = writer.sheets['Maiores Exposições']
            worksheet.set_column('D5:E24', 12, percent)
            worksheet.set_column('C5:C24', 12, numero_float)

            # 20 maiores exposiçoes por produto
            prod=tp_isin_mtm[['contraparte','isin' , 'tipo_prod', 'mtm_info', 'perda_esperada']]
            prod=prod.groupby(['contraparte', 'isin', 'tipo_prod'], as_index=False).sum()
            prod['perc_credito']=prod['mtm_info']/pl_credito
            prod['perc_pe']=prod['perda_esperada']/prod['mtm_info']
            prod=prod.sort(['mtm_info'], ascending=False)
            prod=prod.reset_index(level=None, drop=False, inplace=False, col_level=0, col_fill='')

            del prod['perda_esperada']
            del prod['index']

            prod2=prod[0:20]
            prod2.to_excel(writer,index=False, sheet_name='Maiores Exposições', startrow =28, startcol=1, header =['Contraparte','Isin', 'Tipo de Produto','Exposição', '% PL de crédito', '% Perda Esperada'])

            worksheet = writer.sheets['Maiores Exposições']
            worksheet.set_column('F32:G51', 12, percent)
            worksheet.set_column('E32:E51', 12, numero_float)

            writer.save()

            # inserir caracteristicas
            wb = openpyxl.load_workbook(end+cnpj+"_"+dt_base+' relatorio_credito.xlsx')

            #####FORMATOS
            #Fonte
            fontObj1=Font(name='Calibri', bold=True, size =24,color='404040')
            fontObj2=Font(name='Calibri', bold=False, size =11,color='404040')

            #Borda
            borderObj1=Border(bottom=Side(border_style='double'),top=Side(border_style='thin'))
            borderObj2=Border()

            #Cor
            colorObj1=PatternFill(patternType='solid', fgColor=Color('FFE600'))

            #Alinhamento
            alinObj1=Alignment(vertical='center',horizontal='center')
            alinObj2=Alignment(vertical='center',horizontal='left')
            alinObj3=Alignment(vertical='center',horizontal='right')

            #####Dados institucionais

            #    #####Resumo
            sheet1=wb.get_sheet_by_name('Dados Institucionais')

            #Retira as gridlines
            sheet1.sheet_view.showGridLines = False

            #Formatação tamanho das linhas
            sheet1.row_dimensions[1].height = 90

            #Formata cor da fonte de todas as células
            for row in sheet1.range('B2:C20'):
                for cell in row:
                    cell.font=fontObj2

            #Formata o título
            sheet1.merge_cells('B2:C2')
            sheet1['B2']='Dados Institucionais'
            sheet1['B2'].font=fontObj1
            sheet1['B2'].alignment=alinObj2

            for row in sheet1.range('B2:C2'):
                for cell in row:
                    cell.border=borderObj1

            #Cria a parte de informações institucionais e resumo do relatório de crédito
            sheet1['B4']='Nome'
            sheet1['C4']=header_nome
            sheet1['B5']='CNPJ'
            sheet1['C5']=cnpj_fundo
            sheet1['B6']='Administrador'
            sheet1['C6']=administrador
            sheet1['B7']='Gestor'
            sheet1['C7']=gestor

            #Formatação tamanho das colunas
            sheet1.column_dimensions['A'].width = 2
            sheet1.column_dimensions['B'].width = 15
            sheet1.column_dimensions['C'].width = 100

            ####Resumo
            sheet1=wb.get_sheet_by_name('Resumo')

            #Retira as gridlines
            sheet1.sheet_view.showGridLines = False

            #Formatação tamanho das linhas
            sheet1.row_dimensions[1].height = 90

            #Formata cor da fonte de todas as células
            for row in sheet1.range('B2:F100'):
                for cell in row:
                    cell.font=fontObj2

            #Formata o título
            sheet1.merge_cells('B2:F2')
            sheet1['B2']='Relatório Quantitativo de Risco de Crédito'
            sheet1['B2'].font=fontObj1

            for row in sheet1.range('B2:F2'):
                for cell in row:
                    cell.border=borderObj1

            sheet1['B4']='Nível de Confiança'
            sheet1['C4']=percentil
            sheet1['B5']='PL do fundo (R$)'
            sheet1['C5']=pl_info
            sheet1['B6']='PL de crédito (R$)'
            sheet1['C6']=pl_credito
            sheet1['B7']='Duration - Crédito (DU)'
            sheet1['C7']=duration_carteira
            sheet1['B8']='% de crédito privado'
            sheet1['C8']=pl_credito/pl_info
            sheet1['B9']='Perda Esperada (em relação ao PL crédito)'
            sheet1['C9']=pe
            sheet1['B10']='Percentil da Perda ('+str(percentil)+')'
            sheet1['C10']=percentil_conf
            sheet1['B11']='Rating médio da carteira'
            sheet1['C11']=rtg_lp
            sheet1['B12']='Perda Esperada (em relação ao PL)'
            sheet1['C12']=pe_carteira

            #Formata os headers da tabela
            for row in sheet1.range('B15:F15'):
                for cell in row:
                    cell.fill=colorObj1
                    cell.border=borderObj2

            #Formata os headers da tabela
            for row in sheet1.range('B26:F26'):
                for cell in row:
                    cell.fill=colorObj1
                    cell.border=borderObj2

            #Formata os formatos de número
            for row in sheet1.range('C8:C10'):
                for cell in row:
                    cell.number_format = "0.0%"
                    cell.alignment=alinObj3

            sheet1['C4'].number_format = "0.00"

            #Formatação tamanho das colunas
            sheet1.column_dimensions['A'].width = 2
            sheet1.column_dimensions['B'].width = 40
            sheet1.column_dimensions['C'].width = 25
            sheet1.column_dimensions['D'].width = 25
            sheet1.column_dimensions['E'].width = 25
            sheet1.column_dimensions['F'].width = 25

            ####Maiores exposições
            sheet1=wb.get_sheet_by_name('Maiores Exposições')

            #Retira as gridlines
            sheet1.sheet_view.showGridLines = False

            #Formatação tamanho das linhas
            sheet1.row_dimensions[1].height = 90

            #Formata cor da fonte de todas as células
            for row in sheet1.range('B2:G100'):
                for cell in row:
                    cell.font=fontObj2

            #Formata o título
            sheet1.merge_cells('B2:G2')
            sheet1['B2']='20 Maiores Exposições por Contraparte'
            sheet1['B2'].font=fontObj1

            for row in sheet1.range('B2:G2'):
                for cell in row:
                    cell.border=borderObj1

            #Formata os headers da tabela
            for row in sheet1.range('B4:G4'):
                for cell in row:
                    cell.fill=colorObj1
                    cell.border=borderObj2

            #Formata os headers da tabela
            for row in sheet1.range('B29:G29'):
                for cell in row:
                    cell.fill=colorObj1
                    cell.border=borderObj2

            #Formata os formatos de número
            for row in sheet1.range('E5:E20'):
                for cell in row:
                    cell.number_format = "0.0%"

            #Formatação tamanho das colunas
            sheet1.column_dimensions['A'].width = 2
            sheet1.column_dimensions['B'].width = 100
            sheet1.column_dimensions['C'].width = 20
            sheet1.column_dimensions['D'].width = 20
            sheet1.column_dimensions['E'].width = 20
            sheet1.column_dimensions['F'].width = 20
            sheet1.column_dimensions['G'].width = 20

            wb.save(end+cnpj+"_"+dt_base+' relatorio_credito.xlsx')

        else:
            cnpj_sem_credito = pd.DataFrame(columns=['cnpj','id_relatorio_qo','data_bd'])
            cnpj_sem_credito['cnpj'] = cnpj
            cnpj_sem_credito['id_relatorio_qo'] = id_relatorio_qo_str
            cnpj_sem_credito['data_bd'] = horario_bd
            pd.io.sql.to_sql(cnpj_sem_credito, name='rel_credito_log', con=connection,if_exists="append", flavor='mysql', index=0, chunksize=5000)

    #Busca lista dos fundos de primeiro nivel na carteira da HDI
    query='select * from projeto_inv.xml_header where cnpjcpf="' + cnpj_hdi +'" and dtposicao='+'"'+dt_base+'";'
    df=pd.read_sql(query, con=connection)

    if len(df)==0:
        x='select * from projeto_inv.xml_header where cnpj="' + cnpj_hdi +'" and dtposicao='+'"'+dt_base+'";'
        df=pd.read_sql(x, con=connection)

    df=df.sort(['cnpj', 'cnpjcpf','data_bd'], ascending=[True, True, False])
    df=df.drop_duplicates(subset=['cnpj', 'cnpjcpf'], take_last=False)
    df=df.reset_index(level=None, drop=False, inplace=False, col_level=0, col_fill='')
    del df['index']

    header_id_carteira_fundos=df.get_value(0,'header_id').astype(str) #Utiliza o header da carteira da HDI como chave para a query da lista

    lista_query='SELECT cnpj from projeto_inv.lista_fundos where data_bd=(select max(data_bd) from projeto_inv.lista_fundos where header_id="'+header_id_carteira_fundos+'");'
    lista_cnpj=pd.read_sql(lista_query, con=connection)
    lista=lista_cnpj['cnpj'].tolist()

    horario_bd = datetime.datetime.today()

    for cnpj in lista:
        relatorio_credito(dt_base,cnpj,nivel_confianca,horario_bd)